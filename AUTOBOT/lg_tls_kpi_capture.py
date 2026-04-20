# %%
import os
import io
import time
from datetime import datetime

from PIL import Image
import win32clipboard

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, NoAlertPresentException

from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse

# %%
expedia_url = "https://console.vap.expedia.com/analytics-console-user-interface/reports/personal_reports/lg_tls_hcm_kpi_2025"
teams_url = "https://teams.microsoft.com/"
group_chat_name = "Expedia - Only Team Leaders"
parsed_url = urlparse(expedia_url)
path_fragment = parsed_url.path.split('/')[-1]
report_name = "Updated LG HCM KPI"
wait_seconds = 300
start_time = datetime.now()

chrome_options = Options()
chrome_options.add_argument(r'--user-data-dir=C:/temp/new_chrome_profile')
chrome_options.add_argument(r'--profile-directory=Default')
chrome_options.add_argument("--start-maximized")

service = Service(ChromeDriverManager().install())


service = Service(r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\chromedriver-win64\chromedriver.exe") 
driver = webdriver.Chrome(service=service, options=chrome_options)


if not os.path.exists('screenshots'):
    os.makedirs('screenshots')
def send_to_clipboard(image):
    output = io.BytesIO()
    image.convert("RGB").save(output, "BMP")
    data = output.getvalue()[14:]
    output.close()

    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
    win32clipboard.CloseClipboard()
    print("Image copied to clipboard!")
    
wait = WebDriverWait(driver, 15)
driver.get(expedia_url)

def check_and_login(driver, expedia_url, wait_time=10):
    driver.get(expedia_url)
    time.sleep(10)  # Let the page load

    try:
        sign_in_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="console-okta-sign-in"]'))
        )
        print("🔑 Sign-in required detected! Clicking...")
        sign_in_button.click()
        time.sleep(2)
        try:
            keep_signed_in_label = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="input36"][data-se-for-name="rememberMe"]'))
            )
            keep_signed_in_label.click()
            time.sleep(1)
        except TimeoutException:
            print("No 'Keep me signed in' option found. Skipping.")
        try:
            next_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'input.button.button-primary[type="submit"][value="Next"]'))
            )
            next_button.click()
            time.sleep(10)
        except TimeoutException:
            print("No 'Next' button found. Skipping.")
        try:
            alert = driver.switch_to.alert
            alert.accept()  # Bấm nút OK trên alert
        except NoAlertPresentException:
            print("No 'Alert' button found. Skipping.")
            
        print("🎉 Login successful! Reloading the page...")
        driver.get(expedia_url)
    except TimeoutException:
        print("✅ No sign-in required. Continuing with expedia_url...")

check_and_login(driver, expedia_url)

wait.until(EC.url_contains(path_fragment))
time.sleep(wait_seconds)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
screenshot_path = f'screenshots/screenshot_{timestamp}.png'

driver.save_screenshot(screenshot_path)
print(f"Screenshot saved at: {screenshot_path}")

image = Image.open(screenshot_path)

send_to_clipboard(image)

driver.get(teams_url)

import datetime
import time

try:
    wait = WebDriverWait(driver, 10)
    group_chat = wait.until(EC.presence_of_element_located((By.XPATH, f"//span[contains(text(),'{group_chat_name}')]")))
    group_chat.click()
except TimeoutException:
    pass

time.sleep(3)

chat_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div[role='textbox']")))
chat_box.click()

now = datetime.datetime.now()
# ist = now - datetime.timedelta(hours=1.5)
realtime = now.strftime("%I:%M %p")  
chat_text = f"{report_name} updated at {realtime} (VNT)"

actions = ActionChains(driver)
actions.send_keys(chat_text)
actions.perform()

time.sleep(2)

actions = ActionChains(driver)
actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()

time.sleep(2)

actions.send_keys(Keys.ENTER).perform()

time.sleep(10)

driver.quit()

# %%
from openpyxl import Workbook, load_workbook
from datetime import datetime

end_time = datetime.now()
execution_time = (end_time - start_time).total_seconds()

log_file = "bot_log/bot_capture_log.xlsx"

if not os.path.exists(log_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["BOT Name", "Run at", "End at", "Execution Time (seconds)"])
else:
    wb = load_workbook(log_file)
    ws = wb["Log"]

ws.append([
    report_name,
    start_time.strftime('%Y-%m-%d %H:%M:%S'),
    end_time.strftime('%Y-%m-%d %H:%M:%S'),
    execution_time
])

wb.save(log_file)


