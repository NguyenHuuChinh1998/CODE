# %%
import os
import re
import datetime as _dt
import glob as glob_module
import polars as pl
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from IPython.display import display

# ── Path Setup ─────────────────────────────────────────────────────────────────
first_glob = os.path.expanduser('~').replace('\\', '/')
test_path  = f'{first_glob}/Concentrix Corporation'
if not os.path.exists(test_path):
    raise FileNotFoundError(f'Not found the path: {test_path}')

folder_paths = {
    'input_ou_mail'       : f'{first_glob}/Concentrix Corporation/WFM-Expedia-HCM - Branding files/Rawdata/INPUT_OU_MAIL',
    'output_ou_mail'      : f'{first_glob}/Concentrix Corporation/WFM-Expedia-HCM - Branding files/BI_Task/CODE/Resources/ou_email.csv',
    'output_ou_details'   : f'{first_glob}/Concentrix Corporation/WFM-Expedia-HCM - Branding files/BI_Task/CODE/Resources/ou_details.csv',
    'iex_intervals_output': f'{first_glob}/Concentrix Corporation/WFM-Expedia-HCM - Branding files/Rawdata/OUTPUT_AGENT_IEX_INTERVALS',
    'hc_extend_by_month'  : f'{first_glob}/Concentrix Corporation/WFM-Expedia-HCM - Branding files/Headcount/HC Extend by Month',
}

output_ou_mail    = folder_paths['output_ou_mail']
output_ou_details = folder_paths['output_ou_details']

# ── Config ─────────────────────────────────────────────────────────────────────
ROW_START = 5
ROW_END   = 53

LG_SHEET_CONFIG = {
    'VNM - OU'  : ('E', 'Y'),
    'Kol - OU'  : ('E', 'Y'),
    'Pune - OU' : ('E', 'Y'),
    'Gobal - OU': ('G', 'AA'),
    'Egypt - OU': ('E', 'Y'),
}
NL_SHEET_CONFIG = {
    'VNM - OU'  : ('E', 'Y'),
    'Kol - OU'  : ('E', 'Y'),
    'Pune - OU' : ('E', 'Y'),
    'Gobal - OU': ('F', 'Z'),
    'Egypt - OU': ('E', 'Y'),
}

DAYS            = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
METRICS         = ['Req W', 'Prov', 'DIFF']
DAY_METRIC_COLS = [f'{d}_{m}' for d in DAYS for m in METRICS]
DAY_OFFSET      = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}

TZ_OFFSETS = {
    'VNT': _dt.timedelta(hours=15),
    'IST': _dt.timedelta(hours=13, minutes=30),
    'CLT': _dt.timedelta(hours=4),
}

OU_TO_SITE = {
    'VNM - OU'   : 'Concentrix (Ho Chi Minh City)',
    'Kol - OU'   : 'Concentrix (Kolkata)',
    'Pune - OU'  : 'Concentrix (Pune)',
    'Gobal - OU' : 'Concentrix (Global)',
    'Global - OU': 'Concentrix (Global)',
    'Egypt - OU' : 'Concentrix (Cairo)',
}

LOB_MAPPING = {
    'GEN_GEN_EN_GCS_GLG_CHT'           : 'Lodging chat',
    'GEN_GEN_EN_GCS_GNL_CHT'           : 'Non-Lodging chat',
    'GEN_GEN_EN_GCS_GLG_CHT_Concentrix': 'Lodging chat',
    'GEN_GEN_EN_GCS_GNL_CHT_Concentrix': 'Non-Lodging chat',
    'Lodging'                           : 'Lodging chat',
    'Lodging_Nesting'                   : 'Lodging chat',
    'Non_Lodging'                       : 'Non-Lodging chat',
}

LEAVE_LIST = ['Unscheduled', 'PTO', 'Termination', 'Offline', 'Paid Leave']

# ── Helpers ────────────────────────────────────────────────────────────────────
def find_latest_file(folder: str, *keywords: str) -> str:
    matches = []
    for kw in keywords:
        matches += glob_module.glob(os.path.join(folder, f'*{kw}*.xlsx'))
    matches = list(set(matches))
    if not matches:
        raise FileNotFoundError(
            f'No file matched any of {keywords} in: {folder}')
    def _file_date_key(path: str) -> _dt.date:
        d = extract_week_start(path)
        return d if d else _dt.date.fromtimestamp(os.path.getmtime(path))
    matches.sort(key=_file_date_key, reverse=True)
    for path in matches:
        try:
            with open(path, 'rb') as f:
                f.read(4)
            print(f'  [FOUND] {os.path.basename(path)} '
                  f'(week={_file_date_key(path)}, from {len(matches)} candidate(s))')
            return path
        except (IOError, OSError):
            print(f'  [SKIP]  {os.path.basename(path)} '
                  f'— OneDrive cloud-only, not downloaded yet')
    names = [os.path.basename(p) for p in matches]
    raise FileNotFoundError(
        f'No locally available file for {keywords}.\n'
        f'  Files found (cloud-only): {names}\n'
        f'  Fix: right-click in Explorer -> "Always keep on this device"'
    )

def extract_week_start(file_path: str) -> _dt.date | None:
    match = re.search(r'(\d{2})_(\d{2})_(\d{2})', os.path.basename(file_path))
    if not match:
        return None
    month, day, year = int(match[1]), int(match[2]), int(match[3]) + 2000
    return _dt.date(year, month, day)

def _resolve_sheet(sheetnames: list, target: str) -> str | None:
    if target in sheetnames:
        return target
    def _norm(s): return s.lower().replace(' ', '').replace('-', '')
    t_norm = _norm(target)
    for s in sheetnames:
        if _norm(s) == t_norm:
            return s
    t_loose = t_norm.replace('global', 'gobal')
    for s in sheetnames:
        if _norm(s).replace('global', 'gobal') == t_loose:
            return s
    return None

def _map_site(ou_val: str) -> str:
    if ou_val in OU_TO_SITE:
        return OU_TO_SITE[ou_val]
    norm = lambda s: s.lower().replace(' ', '').replace('-', '')
    for k, v in OU_TO_SITE.items():
        if norm(k).replace('global', 'gobal') == norm(ou_val).replace('global', 'gobal'):
            return v
    return ou_val

def _to_time_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, _dt.time):
        return val.strftime('%H:%M')
    if isinstance(val, _dt.datetime):
        return val.strftime('%H:%M')
    if isinstance(val, str):
        for fmt in ('%H:%M:%S', '%H:%M'):
            try:
                return _dt.datetime.strptime(val.strip(), fmt).strftime('%H:%M')
            except ValueError:
                continue
        return val
    return None

def input_data(folder_path: str) -> pl.DataFrame:
    file_paths = (glob_module.glob(f'{folder_path}/*.xlsx') +
                  glob_module.glob(f'{folder_path}/*.csv'))
    df_list = []
    for file in file_paths:
        basename = os.path.basename(file)
        match = re.match(r'^(\d{4})', basename)
        if match and int(match.group(1)) < 2026:
            continue
        if file.endswith('.xlsx'):
            df = pl.read_excel(file)
        elif file.endswith('.csv'):
            try:
                df = pl.read_csv(file, encoding='utf-8')
            except Exception:
                df = pl.read_csv(file, encoding='ISO-8859-1', ignore_errors=True)
        df = df.with_columns(pl.all().cast(pl.String))
        df_list.append(df)
    return pl.concat(df_list, how='vertical') if df_list else pl.DataFrame()

def read_ou_sheet(ws, col_start: str, col_end: str,
                  label_col: str = 'A',
                  row_start: int = ROW_START,
                  row_end: int   = ROW_END) -> pl.DataFrame:
    lbl_idx       = column_index_from_string(label_col)
    d_start       = column_index_from_string(col_start)
    d_end         = column_index_from_string(col_end)
    tmp_data_cols = [get_column_letter(c) for c in range(d_start, d_end + 1)]
    pst_list      = []
    data_lists    = {col: [] for col in tmp_data_cols}
    for row in range(row_start, row_end + 1):
        pst_list.append(_to_time_str(ws.cell(row=row, column=lbl_idx).value))
        for i, c in enumerate(range(d_start, d_end + 1)):
            v = ws.cell(row=row, column=c).value
            try:
                data_lists[tmp_data_cols[i]].append(
                    float(v) if v is not None else None)
            except (TypeError, ValueError):
                data_lists[tmp_data_cols[i]].append(None)
    df = pl.DataFrame({'PST': pst_list, **data_lists})
    df = df.filter(
        pl.col('PST').is_not_null() &
        (pl.col('PST').str.strip_chars() != 'PST'))
    rename_map = {
        old: new for old, new in
        zip(tmp_data_cols, DAY_METRIC_COLS[:len(tmp_data_cols)])
    }
    return df.rename(rename_map)

def load_ou_file(file_path: str, sheet_config: dict, file_label: str) -> dict:
    wb      = openpyxl.load_workbook(file_path, data_only=True)
    week_dt = extract_week_start(file_path)
    print(f'\n📗 {os.path.basename(file_path)}  |  Week start: {week_dt}')
    print(f'   Sheets: {wb.sheetnames}')
    result = {}
    for target, (col_s, col_e) in sheet_config.items():
        actual = _resolve_sheet(wb.sheetnames, target)
        if actual is None:
            print(f'   ⚠️  Sheet not found: "{target}" — skipped')
            continue
        df = read_ou_sheet(wb[actual], col_s, col_e).with_columns([
            pl.lit(week_dt).alias('Week'),
            pl.lit(file_label).alias('LOB'),
            pl.lit(_map_site(actual)).alias('Site'),
        ])
        result[actual] = df
        print(f'   ✅ "{actual}" → {df.shape}')
    wb.close()
    return result

# ── Load OU Mail ───────────────────────────────────────────────────────────────
lg_path = find_latest_file(folder_paths['input_ou_mail'], 'Global OU LG Chat')
nl_path = find_latest_file(folder_paths['input_ou_mail'], 'Global OU NL Chat')

lg_data = load_ou_file(lg_path, LG_SHEET_CONFIG, file_label='Lodging chat')
nl_data = load_ou_file(nl_path, NL_SHEET_CONFIG, file_label='Non-Lodging chat')

# ── Combine & Unpivot ──────────────────────────────────────────────────────────
all_combined = pl.concat([*lg_data.values(), *nl_data.values()], how='vertical')
id_cols  = ['Site', 'LOB', 'Week', 'PST']
val_cols = [c for c in all_combined.columns if c not in id_cols]

df_long = (
    all_combined
    .unpivot(index=id_cols, on=val_cols,
             variable_name='_col', value_name='Value')
    .with_columns([
        pl.col('_col').str.split_exact('_', 1).struct.field('field_0').alias('Day'),
        pl.col('_col').str.split_exact('_', 1).struct.field('field_1').alias('OU Status'),
    ])
    .with_columns(
        pl.col('Day').replace_strict(DAY_OFFSET, return_dtype=pl.Int32).alias('_offset'))
    .with_columns(
        (pl.col('Week').cast(pl.Datetime('us')) +
         pl.duration(days=pl.col('_offset')))
        .dt.date().alias('PST Date')
    )
    .with_columns(
        pl.concat_str([
            pl.col('PST Date').cast(pl.String), pl.lit(' '), pl.col('PST')
        ])
        .str.to_datetime('%Y-%m-%d %H:%M').alias('PST Datetime')
    )
)

for tz, td in TZ_OFFSETS.items():
    us = int(td.total_seconds() * 1_000_000)
    df_long = (
        df_long
        .with_columns(
            (pl.col('PST Datetime') + pl.duration(microseconds=us))
            .alias(f'{tz} Datetime')
        )
        .with_columns(
            pl.col(f'{tz} Datetime').dt.date().alias(f'{tz} Date')
        )
    )

df_long = (
    df_long
    .with_columns([
        pl.col('Value').round(2),
        pl.col('VNT Datetime').dt.strftime('%H:%M').alias('VNT'),
    ])
    .drop(['_col', 'Day', '_offset'])
    .select([
        'Site', 'LOB', 'Week',
        'PST Date', 'PST', 'PST Datetime',
        'VNT Datetime', 'VNT Date', 'VNT',
        'IST Datetime', 'IST Date',
        'CLT Datetime', 'CLT Date',
        'OU Status', 'Value',
    ])
)

# ── Export df_long ─────────────────────────────────────────────────────────────
os.makedirs(os.path.dirname(output_ou_mail), exist_ok=True)
df_long.write_csv(output_ou_mail)

print(f'\nLong format : {df_long.shape}')
print(f'Columns     : {df_long.columns}')
display(df_long.head(21))

# ── IC_HCM_Details_Log ─────────────────────────────────────────────────────────
IEX_Intervals_Input = (
    input_data(folder_paths['iex_intervals_output'])
    .with_columns([
        pl.col('OracleID').cast(pl.Int64),
        pl.col('IEX ID').cast(pl.Int64),
        pl.col('Week_Monday').str.to_date('%Y-%m-%d'),
        pl.col('Date_Converted').str.to_date('%Y-%m-%d'),
        pl.col('VNT_Intervals').str.to_datetime('%Y-%m-%d %H:%M:%S'),
        pl.col('PST_Intervals').str.to_datetime('%Y-%m-%d %H:%M:%S'),
        pl.col('Datetime_Start_Time').str.to_datetime('%Y-%m-%d %H:%M:%S'),
        pl.col('Datetime_End_Time').str.to_datetime('%Y-%m-%d %H:%M:%S'),
        pl.col('Duration').cast(pl.Float64),
    ])
    .filter(pl.col('Date_Converted').dt.year() == 2026)
)

HC_EXTEND_COMBINED = (
    input_data(folder_paths['hc_extend_by_month'])
    .filter(pl.col('Year') == '2026')
    .select([
        pl.col('Date').str.to_date('%Y-%m-%d'),
        (pl.col('Month') + '-01').str.to_date('%b-%y-%d').dt.strftime('%y_%m').alias('Month'),
        pl.col('Email Id'), pl.col('Alias'), pl.col('Designation'),
        pl.col('Supervisor Name'), pl.col('LOB'), pl.col('Active'), 'Wave',
    ])
)

IC_HCM_Details_Log = (
    IEX_Intervals_Input
    .join(
        HC_EXTEND_COMBINED.select(['Date', 'Email Id', 'LOB']),
        left_on=['Date_Converted', 'Email Id'],
        right_on=['Date', 'Email Id'],
        how='left',
    )
    .with_columns(
        pl.col('LOB').replace_strict(LOB_MAPPING, default=pl.col('LOB')))
    .group_by(['LOB', 'VNT_Intervals', 'PST_Intervals']).agg([
        (pl.col('Duration')
            .filter(pl.col('Scheduled Activity').is_in(['Open Time', 'Extra Hours']))
            .sum() * 2).alias('Scheduled_Open_Time'),
        (pl.col('Duration')
            .filter(pl.col('Scheduled Activity').str.contains('Break'))
            .sum() * 2).alias('Scheduled_Break'),
        (pl.col('Duration')
            .filter(pl.col('Scheduled Activity').str.contains('Lunch'))
            .sum() * 2).alias('Scheduled_Lunch'),
        (pl.col('Duration')
            .filter(pl.col('Scheduled Activity').str.contains('Training|Coaching'))
            .sum() * 2).alias('Scheduled_Training/Coaching'),
        (pl.col('Duration')
            .filter(pl.col('Scheduled Activity').is_in(LEAVE_LIST))
            .sum() * 2).alias('Scheduled_Leave'),
        (pl.col('Duration')
            .filter(pl.col('Scheduled Activity') == 'No Call/No Show')
            .sum() * 2).alias('Scheduled_NCNS'),
        (pl.col('Duration')
            .filter(pl.col('Scheduled Activity') == 'Termination')
            .sum() * 2).alias('Scheduled_Terminated'),
    ])
    .with_columns([
        pl.col('PST_Intervals').dt.strftime('%Y-%m').alias('PST_Month'),
        pl.col('PST_Intervals').dt.date().alias('PST_Date'),
        pl.concat_str([
            pl.col('PST_Intervals').dt.strftime('%H:%M'), pl.lit('-'),
            (pl.col('PST_Intervals') + pl.duration(minutes=30)).dt.strftime('%H:%M'),
        ]).alias('PST_Interval_Range'),
        pl.col('VNT_Intervals').dt.date().alias('VNT_Date'),
        pl.concat_str([
            pl.col('VNT_Intervals').dt.strftime('%H:%M'), pl.lit('-'),
            (pl.col('VNT_Intervals') + pl.duration(minutes=30)).dt.strftime('%H:%M'),
        ]).alias('VNT_Interval_Range'),
    ])
    .select([
        'LOB', 'PST_Month', 'PST_Date', 'PST_Intervals', 'PST_Interval_Range',
        'VNT_Date', 'VNT_Intervals', 'VNT_Interval_Range',
        'Scheduled_Open_Time', 'Scheduled_Break', 'Scheduled_Lunch',
        'Scheduled_Training/Coaching', 'Scheduled_Leave',
        'Scheduled_NCNS', 'Scheduled_Terminated',
    ])
)

print(f'✅ IC_HCM_Details_Log : {IC_HCM_Details_Log.shape}')

# ── df_hcm_wide (HCM + Global) ────────────────────────────────────────────────
df_hcm_wide = (
    df_long
    .filter(pl.col('Site') == 'Concentrix (Ho Chi Minh City)')
    .select(['LOB', 'PST Datetime', 'OU Status', 'Value'])
    .pivot(on='OU Status', index=['LOB', 'PST Datetime'], values='Value')
    .rename({'Req W': 'Req', 'DIFF': 'Diff'})
)

df_global_wide = (
    df_long
    .filter(pl.col('Site') == 'Concentrix (Global)')
    .select(['LOB', 'PST Datetime', 'OU Status', 'Value'])
    .pivot(on='OU Status', index=['LOB', 'PST Datetime'], values='Value')
    .rename({'Req W': 'Global_Req', 'Prov': 'Global_Prov', 'DIFF': 'Global_Diff'})
)

df_hcm_wide = df_hcm_wide.join(df_global_wide, on=['LOB', 'PST Datetime'], how='left')

# ── Final Merge ────────────────────────────────────────────────────────────────
df_merged = (
    df_hcm_wide
    .join(
        IC_HCM_Details_Log,
        left_on=['LOB', 'PST Datetime'],
        right_on=['LOB', 'PST_Intervals'],
        how='left',
    )
    .rename({'Req': 'VN_Req', 'Prov': 'VN_Prov', 'Diff': 'VN_Diff'})
    .select([
        'LOB', 'PST Datetime',
        'PST_Month', 'PST_Date', 'PST_Interval_Range',
        'VNT_Date', 'VNT_Intervals', 'VNT_Interval_Range',
        'VN_Req', 'VN_Prov', 'VN_Diff',
        'Global_Req', 'Global_Prov', 'Global_Diff',
        'Scheduled_Open_Time', 'Scheduled_Break', 'Scheduled_Lunch',
        'Scheduled_Training/Coaching', 'Scheduled_Leave',
        'Scheduled_NCNS', 'Scheduled_Terminated',
    ])
)

os.makedirs(os.path.dirname(output_ou_details), exist_ok=True)
df_merged.write_csv(output_ou_details)

print(f'\ndf_long     : {df_long.shape}')
print(f'df_hcm_wide : {df_hcm_wide.shape}')
print(f'df_merged   : {df_merged.shape}')
print(f'Columns     : {df_merged.columns}')
display(df_merged.head(10))

# %%
import json, requests
import pandas as pd
from datetime import datetime, timedelta

TEAMS_WEBHOOK_URL = "https://default599e51d62f8c43478e591f795a51a9.8c.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/d9dfae822f4941d0be070dd295d55658/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=zQ76qlawVl-CtgQ1Okym9_Vz4rdbSAa0Mc7VHESH3N4"
LOB_RENAME   = {"Lodging chat": "LG Chat", "Non-Lodging chat": "NL Chat"}
LOB_STYLE_OU = {"LG Chat": ("#1565C0","#ffffff"), "NL Chat": ("#6A1B9A","#ffffff")}
IV_STYLE     = {"LG Chat": ("#1E88E5","#ffffff","#64B5F6"),
                "NL Chat": ("#8E24AA","#ffffff","#BA68C8")}
def send_banner():
    ts   = (datetime.utcnow() + timedelta(hours=7)).strftime("%d-%b-%Y  %I:%M %p (VNT)")
    html = (
        '<table cellpadding="0" cellspacing="0" border="0" '
        'style="border-collapse:collapse;width:100%;border-left:5px solid #4A148C;">'
        '<tr>'
        '<td width="5" bgcolor="#4A148C" style="width:5px;">&nbsp;</td>'
        '<td style="padding:8px 14px;">'
        '<span style="font-size:18px;">📈</span>&nbsp;'
        '<b style="color:#4A148C;font-size:16px;">OU Present Report</b><br>'
        '<span style="font-size:11px;opacity:0.75;">'
        f'⏱ <b>{ts}</b>&nbsp;&nbsp;|&nbsp;&nbsp;'
        'Occupancy &amp; staffing overview — Current interval'
        '</span></td></tr></table>'
    )
    try:
        r = requests.post(TEAMS_WEBHOOK_URL, headers={"Content-Type": "application/json"},
                          data=json.dumps({"html": html}), timeout=30)
        print(f"[BANNER] Sent" if r.status_code in (200, 202) else f"[BANNER] Failed [{r.status_code}]")
    except Exception as e:
        print(f"[BANNER] Error: {e}")

now_vnt = datetime.utcnow() + timedelta(hours=7)
now_vnt = datetime.utcnow() + timedelta(hours=7)
min_f   = (now_vnt.minute // 30) * 30
cur_vnt = now_vnt.replace(minute=min_f, second=0, microsecond=0)

df_base = (
    df_merged
    .with_columns(pl.col("LOB").replace(LOB_RENAME))
    .rename({"VNT_Intervals": "VN Intervals", "PST Datetime": "Pacific Intervals"})
    .filter(pl.col("VN Intervals") >= cur_vnt)
    .sort(["LOB", "VN Intervals"])
    .with_row_index("_idx")
    .with_columns((pl.col("_idx") - pl.col("_idx").min().over("LOB")).alias("_rank"))
    .filter(pl.col("_rank") < 10)
    .drop(["_idx", "_rank"])
    .with_columns([
        pl.concat_str([
            pl.col("VN Intervals").dt.strftime("%d-%b"),
            pl.lit("|"),
            pl.col("VN Intervals").dt.strftime("%H:%M"),
        ]).alias("VN Intervals"),
        pl.concat_str([
            pl.col("Pacific Intervals").dt.strftime("%d-%b"),
            pl.lit("|"),
            pl.col("Pacific Intervals").dt.strftime("%H:%M"),
        ]).alias("Pacific Intervals"),
    ])
)

df_t1 = df_base.select([
    "LOB","Pacific Intervals","VN Intervals",
    "VN_Req","VN_Prov","VN_Diff",
    "Global_Req","Global_Prov","Global_Diff","Scheduled_Open_Time",
]).to_pandas()

df_t2 = df_base.select([
    "LOB","Pacific Intervals","VN Intervals","VN_Req","Scheduled_Open_Time",
]).to_pandas()
df_t2["Surplus Heads"] = (df_t2["Scheduled_Open_Time"] - df_t2["VN_Req"]).round(2)

def _surplus_color(v):
    if v >= 5:  return "#1B5E20","#ffffff"
    if v >= 2:  return "#43A047","#ffffff"
    if v >= 0:  return "#A5D6A7","#1a1a1a"
    if v >= -2: return "#FF8F00","#ffffff"
    if v >= -5: return "#E53935","#ffffff"
    return "#B71C1C","#ffffff"

def _heat_color(v, mx, mn=0):
    ratio = max(0.0, min(1.0, (v-mn)/(mx-mn) if (mx-mn)>0 else 0))
    r = int(255 - ratio*(255-102))
    g = int(255 - ratio*(255-187))
    b = int(255 - ratio*(255-106))
    return f"#{r:02x}{g:02x}{b:02x}", "#1a1a1a"

def _interval_cell(v, lv):
    st       = IV_STYLE.get(lv, ("#1E88E5","#ffffff","#64B5F6"))
    color    = st[0]
    parts    = v.split("|", 1)
    date_val = parts[0].strip() if len(parts)==2 else ""
    time_val = parts[1].strip() if len(parts)==2 else v
    return (
        '<td style="padding:4px 8px;border:1px solid #ddd;" nowrap>'
        f'{date_val} | <b style="color:{color};">{time_val}</b>'
        '</td>')

def build_ou_html(df, title, cases=0, summary=''):
    now  = datetime.now()
    sub  = f"Updated {now.strftime('%d-%b-%Y')} · {now.strftime('%I:%M %p')} (VNT)"
    df   = df.copy()
    cols = list(df.columns)

    surplus_idx  = cols.index("Surplus Heads")     if "Surplus Heads"     in cols else -1
    lob_idx      = cols.index("LOB")               if "LOB"               in cols else -1
    vn_iv_idx    = cols.index("VN Intervals")      if "VN Intervals"      in cols else -1
    pac_iv_idx   = cols.index("Pacific Intervals") if "Pacific Intervals" in cols else -1
    interval_idx = {i for i in [vn_iv_idx, pac_iv_idx] if i >= 0}
    diff_cols    = {"VN_Diff","Global_Diff"}
    BAR_COLS     = {"VN_Req","Global_Req","Scheduled_Open_Time"}
    col_max      = {c: pd.to_numeric(df[c],errors='coerce').max() for c in cols if c in BAR_COLS}
    col_min      = {c: pd.to_numeric(df[c],errors='coerce').min() for c in cols if c in BAR_COLS}

    th   = 'bgcolor="#1e3a5f" style="color:#ffffff;padding:7px 12px;border:1px solid #2c4f7c;text-align:center;" nowrap'
    hdrs = [f'<th {th}>{c}</th>' for c in cols]
    if surplus_idx >= 0:
        hdrs.insert(surplus_idx + 1, f'<th {th}>Trend Bar</th>')

    def tdp(bg, v):
        return f'<td bgcolor="{bg}" style="color:#1a1a1a;padding:6px 12px;border:1px solid #ddd;text-align:center;" nowrap>{v}</td>'

    def tdb(bg, fg, v):
        return f'<td bgcolor="{bg}" style="color:{fg};padding:6px 12px;border:1px solid #ddd;font-weight:bold;text-align:center;" nowrap>{v}</td>'

    rows_html = ''
    for i, (_, row) in enumerate(df.iterrows()):
        rbg  = '#f0f4ff' if i % 2 == 0 else '#ffffff'
        lv   = str(row[cols[lob_idx]]) if lob_idx >= 0 else ''
        cells = []
        for j, col in enumerate(cols):
            v = str(row[col]) if pd.notna(row[col]) else '—'

            if j == lob_idx:
                s = LOB_STYLE_OU.get(v)
                cells.append(tdb(s[0], s[1], v) if s else tdp(rbg, v))

            elif j in interval_idx:
                cells.append(_interval_cell(v, lv))

            elif j == surplus_idx:
                try:
                    n         = float(v)
                    bg, fg    = _surplus_color(n)
                    r         = max(-1.0, min(1.0, n / 10))
                    filled    = int(abs(r) * 10)
                    empty     = 10 - filled
                    bar_color = "#43A047" if r >= 0 else "#E53935"
                    bar       = (
                        f'<span style="color:{bar_color};font-family:monospace;'
                        f'letter-spacing:1px;font-weight:bold;">'
                        f'{"█" * filled}{"░" * empty}</span>'
                    )
                    cells.append(
                        f'<td bgcolor="{bg}" style="color:{fg};padding:6px 12px;'
                        f'border:1px solid #ddd;font-weight:bold;text-align:center;" nowrap>'
                        f'{n:+.2f}</td>'
                    )
                    cells.append(
                        f'<td style="padding:6px 12px;border:1px solid #ddd;text-align:center;" nowrap>'
                        f'{bar}</td>'
                    )
                except:
                    cells.append(tdp(rbg, v))
                    cells.append(tdp(rbg, ''))

            elif col in diff_cols:
                try:
                    n = float(v)
                    cells.append(tdb("#1B5E20" if n >= 0 else "#B71C1C", "#ffffff", f"{n:+.2f}"))
                except:
                    cells.append(tdp(rbg, v))

            elif col in BAR_COLS:
                try:
                    n      = float(v)
                    bg, fg = _heat_color(n, col_max.get(col, 1), col_min.get(col, 0))
                    cells.append(tdb(bg, fg, f"{n:.2f}"))
                except:
                    cells.append(tdp(rbg, v))

            else:
                try:
                    cells.append(tdp(rbg, f"{float(v):.2f}"))
                except:
                    cells.append(tdp(rbg, v))

        rows_html += f"<tr>{''.join(cells)}</tr>"

    sh = f'  <span style="font-size:11px;">📊 {summary}</span><br>\n' if summary else ''
    return (
        f'<p>\n  <b style="color:#c0392b;font-size:16px;">🔴 {title}</b><br>\n'
        f'  <span style="font-size:12px;">{sub} &nbsp;|&nbsp; ⚡ <b>{cases} CASES</b></span><br>\n{sh}</p>\n'
        f'<div style="overflow-x:auto;">\n<table border="1" cellpadding="0" cellspacing="0" '
        f'style="border-collapse:collapse;font-size:12px;font-family:Segoe UI,Arial,sans-serif;">\n'
        f'  <thead><tr>{"".join(hdrs)}</tr></thead>\n  <tbody>{rows_html}</tbody>\n</table>\n</div>'
    )

def send_ou_webhook(df, title, cases=0, summary=''):
    payload = {'html': build_ou_html(df, title, cases, summary)}
    try:
        r = requests.post(
            TEAMS_WEBHOOK_URL,
            headers={'Content-Type': 'application/json'},
            data=json.dumps(payload),
            timeout=30
        )
        print(f"✅ Sent: '{title}'" if r.status_code in (200, 202) else f"❌ [{r.status_code}]: {r.text[:200]}")
    except Exception as e:
        print(f"❌ {e}")

send_banner()
win_label = f"{cur_vnt.strftime('%d-%b %H:%M')} → {(cur_vnt + timedelta(minutes=30*7)).strftime('%H:%M')} (VNT)"
send_ou_webhook(df_t1, "OU — VN Staffing Overview",    cases=len(df_t1), summary=win_label)
send_ou_webhook(df_t2, "OU — Surplus / Deficit Heads", cases=len(df_t2), summary=win_label)
print(f"\nWindow: {win_label} | T1:{len(df_t1)} | T2:{len(df_t2)} rows")


