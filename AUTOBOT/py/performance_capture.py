# %%
import os
import io
import time
from datetime import datetime
from urllib.parse import urlparse

from PIL import Image
import win32clipboard

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException

from webdriver_manager.chrome import ChromeDriverManager

# %%
report_name = "Performance"
start_time = datetime.now()

# --- Config ---
expedia_urls = [
    {
        "url": "https://console.vap.expedia.com/analytics-console-user-interface/reports/personal_reports/lg_chat_aht_lc",
        "group_chat": "LG Mini Team Leaders",
        "report_name": "LG Performance"
    },
    {
        "url": "https://console.vap.expedia.com/analytics-console-user-interface/reports/personal_reports/hcm_kpi_trends",
        "group_chat": "Expedia | Non-Lodging Team",
        "report_name": "NL Performance"
    }
]

teams_url = "https://teams.microsoft.com/"
wait_seconds = 240

chrome_options = Options()
chrome_options.add_argument(r'--user-data-dir=C:/temp/new_chrome_profile')
chrome_options.add_argument(r'--profile-directory=Default')
chrome_options.add_argument("--start-maximized")

service = Service(ChromeDriverManager().install())


service = Service(r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\chromedriver-win64\chromedriver.exe") 
driver = webdriver.Chrome(service=service, options=chrome_options)

if not os.path.exists('screenshots'):
    os.makedirs('screenshots')

def send_to_clipboard(image):
    output = io.BytesIO()
    image.convert("RGB").save(output, "BMP")
    data = output.getvalue()[14:]
    output.close()

    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
    win32clipboard.CloseClipboard()

wait = WebDriverWait(driver, 15)

for report in expedia_urls:
    expedia_url = report["url"]
    group_chat_name = report["group_chat"]
    report_name = report["report_name"]
    parsed_url = urlparse(expedia_url)
    path_fragment = parsed_url.path.split('/')[-1]

    driver.get(expedia_url)

    def check_and_login(driver, expedia_url, wait_time=10):
        driver.get(expedia_url)
        time.sleep(10)  # Let the page load

        try:
            sign_in_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="console-okta-sign-in"]'))
            )
            print("🔑 Sign-in required detected! Clicking...")
            sign_in_button.click()
            time.sleep(2)
            try:
                keep_signed_in_label = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="input36"][data-se-for-name="rememberMe"]'))
                )
                keep_signed_in_label.click()
                time.sleep(1)
            except TimeoutException:
                print("No 'Keep me signed in' option found. Skipping.")
            try:
                next_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'input.button.button-primary[type="submit"][value="Next"]'))
                )
                next_button.click()
                time.sleep(10)
            except TimeoutException:
                print("No 'Next' button found. Skipping.")
            print("🎉 Login successful! Reloading the page...")
            driver.get(expedia_url)
        except TimeoutException:
            print("✅ No sign-in required. Continuing with expedia_url...")

    check_and_login(driver, expedia_url)

    wait.until(EC.url_contains(path_fragment))
    time.sleep(wait_seconds)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    screenshot_path = f'screenshots/{report_name.replace(" ", "_")}_{timestamp}.png'
    driver.save_screenshot(screenshot_path)

    image = Image.open(screenshot_path)
    send_to_clipboard(image)

    driver.get(teams_url)

    try:
        group_chat = wait.until(EC.presence_of_element_located((By.XPATH, f"//span[contains(text(),'{group_chat_name}')]")))
        group_chat.click()
    except TimeoutException:
        pass

    time.sleep(3)

    try:
        chat_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div[role='textbox']")))
        chat_box.click()
        realtime = datetime.now().strftime("%I:%M %p")
        chat_text = f"The {report_name} updated at {realtime} (VNT)"

        actions = ActionChains(driver)
        actions.send_keys(chat_text).perform()
        time.sleep(2)
        actions = ActionChains(driver)
        actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
        time.sleep(2)
        actions.send_keys(Keys.ENTER).perform()
        time.sleep(5)
    except TimeoutException:
        pass

# --- Close ---
driver.quit()

from openpyxl import Workbook, load_workbook
from datetime import datetime

end_time = datetime.now()
execution_time = (end_time - start_time).total_seconds()

log_file = "bot_log/bot_capture_log.xlsx"

if not os.path.exists(log_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["BOT Name", "Run at", "End at", "Execution Time (seconds)"])
else:
    wb = load_workbook(log_file)
    ws = wb["Log"]

ws.append([
    report_name,
    start_time.strftime('%Y-%m-%d %H:%M:%S'),
    end_time.strftime('%Y-%m-%d %H:%M:%S'),
    execution_time
])

wb.save(log_file)


