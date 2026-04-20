# %%
import psutil

start_time = datetime.now()
report_name = "Killer BOT"

def kill_process_openconsole(process_name):
    for proc in psutil.process_iter(['name']):
        if proc.info['name'] == process_name:
            print(f"Killing process {proc.pid} - {process_name}")
            proc.kill()


def kill_process_chromedriver(name):
    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
        try:
            if proc.info['name'] and name.lower() in proc.info['name'].lower():
                print(f"Killing process {proc.info['pid']} - {proc.info['name']}")
                proc.kill()
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass

def kill_chrome_with_profile(profile_dir):
    for proc in psutil.process_iter(['pid', 'name', 'cmdline']):
        try:
            if proc.info['name'] and 'chrome.exe' in proc.info['name'].lower():
                if any(profile_dir in arg for arg in proc.info['cmdline']):
                    print(f"Killing Chrome process {proc.info['pid']} with profile {profile_dir}")
                    proc.kill()
        except (psutil.NoSuchProcess, psutil.AccessDenied):
            pass

kill_process_openconsole("OpenConsole.exe")
kill_process_chromedriver("chromedriver.exe")
kill_chrome_with_profile("C:/temp/new_chrome_profile")

from openpyxl import Workbook, load_workbook
from datetime import datetime

end_time = datetime.now()
execution_time = (end_time - start_time).total_seconds()

log_file = "bot_log/bot_capture_log.xlsx"

if not os.path.exists(log_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["BOT Name", "Run at", "End at", "Execution Time (seconds)"])
else:
    wb = load_workbook(log_file)
    ws = wb["Log"]

ws.append([
    report_name,
    start_time.strftime('%Y-%m-%d %H:%M:%S'),
    end_time.strftime('%Y-%m-%d %H:%M:%S'),
    execution_time
])

wb.save(log_file)

