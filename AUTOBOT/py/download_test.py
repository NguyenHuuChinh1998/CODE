# %%
import subprocess
import os
import sys
from datetime import datetime
from openpyxl import Workbook, load_workbook

if hasattr(sys.stdout, 'buffer'):
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception:
        pass

LOG_FILE = (
    r"C:\Users\huuchinh.nguyen\Concentrix Corporation"
    r"\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\bot_log\download_bot_log.xlsx"
)
PROFILE_PATH = r"C:/temp/new_chrome_profile"
os.makedirs(os.path.dirname(LOG_FILE), exist_ok=True)
RUN_ID = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

print(f"\n{'═'*60}")
print(f"  DOWNLOAD BOT  |  Run ID: {RUN_ID}")
print(f"  Log file     :  {os.path.basename(LOG_FILE)}")
print(f"{'═'*60}\n")


def log_step(step, status, message="", duration=0.0):
    headers = ["Run ID", "Timestamp", "Step", "Status", "Message", "Duration (s)"]
    try:
        if os.path.exists(LOG_FILE):
            wb = load_workbook(LOG_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Log"
            ws.append(headers)
        ws.append([
            RUN_ID,
            datetime.now().strftime("%d-%b-%Y %H:%M:%S"),
            step,
            status,
            str(message)[:500],
            round(duration, 1)
        ])
        wb.save(LOG_FILE)
    except Exception as e:
        print(f"  [LOG-FAIL] Cannot write log: {e}")
        import traceback; traceback.print_exc()


def kill_orphan_chromedriver(profile_keyword="new_chrome_profile"):
    print("[CLEANUP] Killing orphan Chrome processes...")

    try:
        r1 = subprocess.run(
            ["taskkill", "/F", "/IM", "chromedriver.exe"],
            capture_output=True, text=True
        )
        ok  = "SUCCESS" in r1.stdout
        msg = "chromedriver.exe terminated" if ok else "No chromedriver.exe process found"
        print(f"  [{'OK' if ok else 'SKIP'}] chromedriver.exe: {msg}")
        log_step("Kill Chromedriver", "✅ OK" if ok else "ℹ️ Skip", msg)
    except Exception as e:
        print(f"  [ERROR] chromedriver: {e}")
        log_step("Kill Chromedriver", "❌ ERROR", str(e))

    try:
        ps_cmd = (
            f"Get-WmiObject Win32_Process -Filter \"name='chrome.exe'\" | "
            f"Where-Object {{ $_.CommandLine -like '*{profile_keyword}*' }} | "
            f"ForEach-Object {{ Stop-Process -Id $_.ProcessId -Force }}"
        )
        r2 = subprocess.run(
            ["powershell", "-NoProfile", "-Command", ps_cmd],
            capture_output=True, text=True
        )
        ok  = r2.returncode == 0
        msg = f"chrome.exe ({profile_keyword}) terminated" if ok else "No matching chrome.exe found"
        print(f"  [{'OK' if ok else 'SKIP'}] chrome.exe profile: {msg}")
        log_step("Kill Chrome", "✅ OK" if ok else "ℹ️ Skip", msg)
    except Exception as e:
        print(f"  [ERROR] chrome.exe: {e}")
        log_step("Kill Chrome", "❌ ERROR", str(e))

    removed = []
    for lock in ["SingletonLock", "SingletonSocket", "SingletonCookie"]:
        lf = os.path.join(PROFILE_PATH, lock)
        try:
            if os.path.exists(lf):
                os.remove(lf)
                removed.append(lock)
        except Exception as e:
            print(f"  [WARN] Cannot remove {lock}: {e}")
            log_step("Remove Lock File", "⚠️ WARN", f"{lock}: {e}")

    if removed:
        msg = f"Removed lock files: {', '.join(removed)}"
        print(f"  [OK] {msg}")
        log_step("Remove Lock Files", "✅ OK", msg)
    else:
        print(f"  [SKIP] No profile lock files found")
        log_step("Remove Lock Files", "ℹ️ Skip", "No lock files present")

    import time; time.sleep(3)
    print("[CLEANUP] Done\n")


log_step("BOT START", "✅ OK", f"Run ID: {RUN_ID}")
kill_orphan_chromedriver()

# %%
import os, time, shutil, glob
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import (
    TimeoutException, NoAlertPresentException, WebDriverException
)
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime, timedelta

CHROMEDRIVER   = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\chromedriver-win64\chromedriver.exe"
SOURCE_FOLDER  = r"C:\temp\expedia_downloads"
BASE_CAPTURE   = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE"
DIRS = {
    "current_agent"        : os.path.join(BASE_CAPTURE, "current_agent"),
    "lc_rawdata"           : os.path.join(BASE_CAPTURE, "lc_rawdata_in_console"),
    "current_interval"     : os.path.join(BASE_CAPTURE, "current_interval"),
    "forecast_realtime"    : os.path.join(BASE_CAPTURE, "forecast_realtime"),
    "forecast_interval_sum": os.path.join(BASE_CAPTURE, "forecast_interval_summary"),
}
for d in DIRS.values():
    os.makedirs(d, exist_ok=True)

URL_BREAKDOWN  = "https://console.vap.expedia.com/analytics-console-user-interface/optics/agentBreakdownRealtimeDashboard"
URL_REALTIME   = "https://console.vap.expedia.com/analytics-console-user-interface/optics/agentRealtime"
URL_FORECAST   = "https://console.vap.expedia.com/analytics-console-user-interface/optics/forecastRealtime"
URL_SHAREPOINT = (
    "https://cnxmail-my.sharepoint.com/shared?listurl=https%3A%2F%2Fcnxmail-my%2E"
    "sharepoint%2Ecom%2Fpersonal%2Fahmed_ahmedkamh_concentrix_com%2FDocuments"
    "&id=%2Fpersonal%2Fahmed_ahmedkamh_concentrix_com%2FDocuments"
)
DST_UCP          = os.path.join(BASE_CAPTURE, "EN- UCP.xlsx")
LOGIN_VERIFY_CSS = "button.settingsButton"
LOGIN_TIMEOUT    = 20
CNX_USER         = "huuchinh.nguyen@concentrix.com"
CNX_PASS         = "Concentrix@130499"


def _ts():
    return datetime.now().strftime("%H:%M:%S")

def _sep(label=""):
    print(f"\n{'─'*60}")
    if label:
        print(f"  {label}")
    print(f"{'─'*60}")


def move_files(keyword, dest_dir, step_label=""):
    moved = 0
    for pat in [
        f"{SOURCE_FOLDER}\\{keyword}*.csv",
        f"{SOURCE_FOLDER}\\{keyword}*.xlsx"
    ]:
        for fp in glob.glob(pat):
            if fp.endswith(".crdownload"):
                continue
            dst = os.path.join(dest_dir, os.path.basename(fp))
            if os.path.exists(dst):
                os.remove(dst)
            shutil.move(fp, dst)
            fname = os.path.basename(fp)
            dest  = os.path.basename(dest_dir)
            print(f"  [MOVE] {fname} -> {dest}/")
            log_step(
                f"Move File{' - ' + step_label if step_label else ''}",
                "✅ OK",
                f"Moved '{fname}' to {dest}/"
            )
            moved += 1
    if not moved:
        msg = f"No file matching '{keyword}*' found in {SOURCE_FOLDER}"
        print(f"  [WARN] {msg}")
        log_step(
            f"Move File{' - ' + step_label if step_label else ''}",
            "⚠️ WARN", msg
        )
    return moved


def dismiss_modal(driver, wait_disappear=True, timeout=10):
    MODAL_CSS = "div.modal__wrapper"
    try:
        modal = driver.find_element(By.CSS_SELECTOR, MODAL_CSS)
        if not modal.is_displayed():
            return
        print(f"  [MODAL] Overlay detected: '{MODAL_CSS}' — attempting to dismiss...")
        log_step("Dismiss Modal", "⏳ Wait",
                 f"Modal '{MODAL_CSS}' is blocking click, attempting dismiss")
        try:
            modal.send_keys(Keys.ESCAPE)
            time.sleep(1)
            print(f"  [MODAL] Sent Escape key")
        except Exception:
            pass
        try:
            ActionChains(driver).move_by_offset(10, 10).click().perform()
            ActionChains(driver).move_by_offset(0, 0).perform()
            time.sleep(1)
            print(f"  [MODAL] Clicked outside modal area")
        except Exception:
            pass
        try:
            driver.execute_script("""
                const btn = document.querySelector(
                    'div.modal__wrapper [aria-label="Close"], ' +
                    'div.modal__wrapper button.close, ' +
                    'div.modal__wrapper .modal__close'
                );
                if (btn) btn.click();
            """)
            time.sleep(1)
            print(f"  [MODAL] JS-clicked close button inside modal")
        except Exception:
            pass
        if wait_disappear:
            try:
                WebDriverWait(driver, timeout).until(
                    EC.invisibility_of_element_located(
                        (By.CSS_SELECTOR, MODAL_CSS)))
                print(f"  [MODAL] Overlay dismissed successfully")
                log_step("Dismiss Modal", "✅ OK", "Modal dismissed, safe to click")
            except TimeoutException:
                print(f"  [MODAL] WARN: Modal still visible after {timeout}s — proceeding anyway")
                log_step("Dismiss Modal", "⚠️ WARN",
                         f"Modal still present after {timeout}s, proceeding anyway")
    except Exception:
        pass


def click_download_csv(driver, wait, keyword=None, step_label="", timeout=30):
    print(f"  [UI] Waiting for dropdown menu to appear...")
    log_step(
        f"UI - Open Menu{' - ' + step_label if step_label else ''}",
        "⏳ Wait",
        "Waiting for uitk-menu-container[aria-hidden=false]"
    )
    wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "div.uitk-menu-container[aria-hidden='false']")))
    print(f"  [UI] Locating 'Download CSV' button...")
    dl_btn = wait.until(EC.element_to_be_clickable((By.XPATH,
        "//div[contains(@class,'uitk-menu-open')][@aria-hidden='false']"
        "//span[text()='Download CSV']/ancestor::button")))
    driver.execute_script("arguments[0].click();", dl_btn)
    log_step(
        f"Click - Download CSV{' - ' + step_label if step_label else ''}",
        "✅ OK",
        f"JS-clicked 'Download CSV' at {_ts()}"
    )
    print(f"  [DOWNLOAD] 'Download CSV' clicked at {_ts()}")
    if keyword:
        print(f"  [DOWNLOAD] Waiting for file '{keyword}*' in {SOURCE_FOLDER}...")
        start = time.time()
        while time.time() - start < timeout:
            matches = [
                f for f in
                glob.glob(f"{SOURCE_FOLDER}\\{keyword}*.csv") +
                glob.glob(f"{SOURCE_FOLDER}\\{keyword}*.xlsx")
                if not f.endswith('.crdownload')
            ]
            if matches:
                time.sleep(0.5)
                elapsed = round(time.time() - start, 1)
                fname   = os.path.basename(matches[0])
                print(f"  [DOWNLOAD] File ready after {elapsed}s: {fname}")
                log_step(
                    f"Download Ready{' - ' + step_label if step_label else ''}",
                    "✅ OK",
                    f"File ready: {fname} ({elapsed}s)"
                )
                return
            time.sleep(0.5)
        msg = f"Timeout {timeout}s — file '{keyword}*' not found"
        print(f"  [WARN] {msg}")
        log_step(
            f"Download Timeout{' - ' + step_label if step_label else ''}",
            "⚠️ WARN", msg
        )
    else:
        print(f"  [DOWNLOAD] No keyword check — waiting 8s for download...")
        log_step(
            f"Download Wait{' - ' + step_label if step_label else ''}",
            "⏳ Wait", "Fixed 8s wait (no keyword)"
        )
        time.sleep(8)


def _has(driver, by, sel):
    try:
        driver.find_element(by, sel)
        return True
    except Exception:
        return False


def _find_clickable_first(driver, selector_list, timeout=4):
    for by, sel in selector_list:
        try:
            return WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, sel)))
        except TimeoutException:
            continue
    return None


def _is_known_login_domain(url: str) -> bool:
    return any(marker in url for marker in [
        "okta.com", "login.microsoftonline.com", "signin.concentrix.com", "/login"
    ])


def _detect_login_stage(driver):
    url = driver.current_url

    if "okta.com" in url:
        return "okta"

    if "login.microsoftonline.com" in url:
        if _has(driver, By.ID, "idBtn_Back"):
            return "ms_stay_signed_in"
        if _has(driver, By.ID, "i0118"):
            return "ms_password"
        if _has(driver, By.ID, "i0116"):
            return "ms_email"
        return None

    if "signin.concentrix.com" in url:
        if _has(driver, By.XPATH,
                "//*[contains(normalize-space(text()),'Choose an authentication method')]"):
            return "cnx_auth_method_chooser"
        return "cnx_passwordless"

    if "vap.expedia.com" in url and "/login" in url:
        return "expedia_login"

    return None


def _handle_expedia_login_page(driver):
    print(f"  [LOGIN] State: expedia_login -> clicking 'Log in with SSO'")
    log_step("Login - Expedia Login Page", "⏳ Wait", "Detected /login page")
    btn_sso = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, 'button[data-testid="console-okta-sign-in"]')))
    btn_sso.click()
    print(f"  [LOGIN] Clicked 'Log in with SSO'")
    log_step("Login - Click SSO", "✅ OK", "SSO button clicked")


def _handle_okta_page(driver):
    print(f"  [LOGIN] State: okta -> entering username")
    log_step("Login - Okta Page", "⏳ Wait", "Detected okta.com sign-in page")

    uf = _find_clickable_first(driver, [
        (By.CSS_SELECTOR, 'input[name="identifier"]'),
        (By.CSS_SELECTOR, 'input[autocomplete="username"]'),
        (By.CSS_SELECTOR, 'input[type="text"]'),
    ], timeout=6)
    if uf is None:
        raise RuntimeError("Okta: username input not found")

    if not (uf.get_attribute('value') or '').strip():
        uf.click()
        uf.send_keys(CNX_USER)
        print(f"  [LOGIN] Entered username: {CNX_USER}")
        log_step("Login - Okta Username", "✅ OK", f"Username: {CNX_USER}")
    time.sleep(0.5)

    try:
        label = WebDriverWait(driver, 4).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'label[data-se-for-name="rememberMe"]')))
        cb_id = label.get_attribute('for')
        cb = driver.find_element(By.ID, cb_id) if cb_id else None
        if cb is not None and not cb.is_selected():
            driver.execute_script("arguments[0].click();", cb)
            print(f"  [LOGIN] Checked 'Keep me signed in' (JS click via label)")
            log_step("Login - Keep Signed In", "✅ OK", "Checkbox JS-clicked via label[data-se-for-name]")
    except TimeoutException:
        try:
            cb = WebDriverWait(driver, 2).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, 'input[name="rememberMe"]')))
            if not cb.is_selected():
                driver.execute_script("arguments[0].click();", cb)
                print(f"  [LOGIN] Checked 'Keep me signed in' (JS click fallback)")
                log_step("Login - Keep Signed In", "✅ OK", "Checkbox JS-clicked via input fallback")
        except TimeoutException:
            print(f"  [LOGIN] Remember me checkbox not found — skipping")
            log_step("Login - Keep Signed In", "ℹ️ Skip", "Checkbox not found")

    next_btn = WebDriverWait(driver, 8).until(EC.element_to_be_clickable(
        (By.CSS_SELECTOR, 'input.button-primary[type="submit"][data-type="save"]')))
    next_btn.click()
    print(f"  [LOGIN] Clicked 'Next' on Okta")
    log_step("Login - Okta Next", "✅ OK", "Next button clicked")


def _handle_cnx_passwordless_page(driver):
    print(f"  [LOGIN] State: cnx_passwordless")
    log_step("Login - CNX Passwordless Page", "⏳ Wait", "Detected signin.concentrix.com")

    try:
        username_tab = WebDriverWait(driver, 4).until(
            EC.element_to_be_clickable((By.XPATH,
                '//*[self::button or self::a]'
                '[contains(translate(normalize-space(text()),'
                '"abcdefghijklmnopqrstuvwxyz","ABCDEFGHIJKLMNOPQRSTUVWXYZ"),"USERNAME")]')))
        username_tab.click()
        time.sleep(0.5)
    except TimeoutException:
        pass

    user_input = WebDriverWait(driver, 8).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, 'input[name="identifier"]')))
    if not (user_input.get_attribute('value') or '').strip():
        user_input.click()
        user_input.send_keys(CNX_USER)
        print(f"  [LOGIN] Entered SSO email: {CNX_USER}")
        log_step("Login - CNX Username", "✅ OK", f"Email: {CNX_USER}")

    try:
        label = WebDriverWait(driver, 4).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, 'label[data-se-for-name="rememberMe"]')))
        checkbox_id = label.get_attribute('for')
        cb = driver.find_element(By.ID, checkbox_id) if checkbox_id else None
        if cb is not None and not cb.is_selected():
            cb.click()
            print(f"  [LOGIN] Checked 'Remember me' (via data-se-for-name)")
            log_step("Login - CNX Remember Me", "✅ OK", "Checkbox checked via data-se-for-name")
    except TimeoutException:
        print(f"  [LOGIN] Remember me checkbox not found — skipping")
        log_step("Login - CNX Remember Me", "ℹ️ Skip", "Checkbox not found")

    next_btn = _find_clickable_first(driver, [
        (By.CSS_SELECTOR, 'button[type="submit"]'),
        (By.CSS_SELECTOR, 'input[type="submit"]'),
        (By.XPATH, '//*[self::button or self::div[@role="button"]][contains(normalize-space(.),"Next")]'),
        (By.XPATH, '//*[self::button or self::div[@role="button"]][contains(normalize-space(.),"Tiếp theo")]'),
    ], timeout=6)
    if next_btn is None:
        raise RuntimeError("CNX passwordless: Next button not found")
    next_btn.click()
    print(f"  [LOGIN] Clicked Next")
    log_step("Login - CNX Next", "✅ OK", "Next button clicked")

    print(f"  [LOGIN] Note: if the account uses passwordless push notification, "
          f"manual approval on a registered device may be required.")
    log_step("Login - CNX Passwordless Note", "ℹ️ Info",
             "May require manual push approval on registered device")


def _handle_cnx_auth_method_chooser(driver):
    print(f"  [LOGIN] State: cnx_auth_method_chooser -> selecting 'Password'")
    log_step("Login - CNX Auth Method Chooser", "⏳ Wait", "Detected method selection screen")

    pwd_option = _find_clickable_first(driver, [
        (By.XPATH,
         "//*[self::button or self::div[@role='button'] or @tabindex]"
         "[.//*[normalize-space(text())='Password']]"
         "[.//*[contains(normalize-space(text()),'Login with Password')]]"),
        (By.XPATH,
         "//*[normalize-space(text())='Password']"
         "/ancestor::*[self::button or self::div[@role='button'] or @tabindex][1]"),
    ], timeout=8)

    if pwd_option is None:
        raise RuntimeError("'Password' option not found on the authentication method selection screen")

    pwd_option.click()
    print(f"  [LOGIN] Selected 'Password' method")
    log_step("Login - CNX Select Password Method", "✅ OK", "Password method selected")


def _handle_ms_email_step(driver):
    print(f"  [LOGIN] State: ms_email")
    log_step("Login - MS Email Step", "⏳ Wait", "Detected AAD email step (#i0116)")
    email_input = WebDriverWait(driver, 8).until(
        EC.presence_of_element_located((By.ID, "i0116")))
    if not (email_input.get_attribute('value') or '').strip():
        email_input.click()
        email_input.send_keys(CNX_USER)
        print(f"  [LOGIN] Entered email: {CNX_USER}")
        log_step("Login - MS Email", "✅ OK", f"Email: {CNX_USER}")
    next_btn = WebDriverWait(driver, 8).until(
        EC.element_to_be_clickable((By.ID, "idSIButton9")))
    next_btn.click()
    print(f"  [LOGIN] Clicked Next (email step)")
    log_step("Login - MS Email Next", "✅ OK", "Next clicked after email")


def _handle_ms_password_step(driver):
    print(f"  [LOGIN] State: ms_password")
    log_step("Login - MS Password Step", "⏳ Wait", "Detected AAD password step (#i0118)")
    pwd_input = WebDriverWait(driver, 8).until(
        EC.presence_of_element_located((By.ID, "i0118")))
    pwd_input.click()
    pwd_input.send_keys(CNX_PASS)
    signin_btn = WebDriverWait(driver, 8).until(
        EC.element_to_be_clickable((By.ID, "idSIButton9")))
    signin_btn.click()
    print(f"  [LOGIN] Entered password and clicked Sign in")
    log_step("Login - MS Password", "✅ OK", "Password entered, sign-in clicked")


def _handle_ms_stay_signed_in(driver):
    print(f"  [LOGIN] State: ms_stay_signed_in")
    log_step("Login - MS Stay Signed In", "⏳ Wait", "Detected KMSI prompt")
    yes_btn = WebDriverWait(driver, 8).until(
        EC.element_to_be_clickable((By.ID, "idSIButton9")))
    yes_btn.click()
    print(f"  [LOGIN] Clicked 'Yes' for Stay signed in")
    log_step("Login - MS Stay Signed In Confirmed", "✅ OK", "Clicked Yes")


LOGIN_STAGE_HANDLERS = {
    "expedia_login":           _handle_expedia_login_page,
    "okta":                    _handle_okta_page,
    "cnx_passwordless":        _handle_cnx_passwordless_page,
    "cnx_auth_method_chooser": _handle_cnx_auth_method_chooser,
    "ms_email":                _handle_ms_email_step,
    "ms_password":             _handle_ms_password_step,
    "ms_stay_signed_in":       _handle_ms_stay_signed_in,
}


def resolve_login_if_needed(driver, is_app_url_fn, max_stage_transitions=14, stage_timeout=25):
    if is_app_url_fn(driver.current_url):
        print(f"  [LOGIN] Session still valid — no login needed")
        log_step("Login - Auth Check", "✅ OK", "Already authenticated, no login needed")
        return True

    print(f"  [LOGIN] Login required detected — starting state-based handling")
    log_step("Login - Required", "ℹ️ Info", f"Redirected to: {driver.current_url[:120]}")

    last_stage = None
    repeat_count = 0

    for _ in range(max_stage_transitions):
        if is_app_url_fn(driver.current_url):
            print(f"  [LOGIN] Login successful | URL: {driver.current_url[:70]}")
            log_step("Login - Success", "✅ OK", f"Authenticated, URL: {driver.current_url[:120]}")
            return True

        stage = _detect_login_stage(driver)

        if stage is None:
            print(f"  [LOGIN] State not yet identified ({driver.current_url[:70]}) — waiting for page to load...")
            time.sleep(2)
            continue

        repeat_count = repeat_count + 1 if stage == last_stage else 0
        last_stage = stage

        if repeat_count >= 3:
            msg = f"State '{stage}' repeated {repeat_count} times with no progress — stopping"
            print(f"  [LOGIN] {msg}")
            log_step(f"Login - {stage} Stuck", "❌ ERROR", msg)
            return False

        handler = LOGIN_STAGE_HANDLERS[stage]
        try:
            handler(driver)
        except Exception as e:
            print(f"  [LOGIN] Error handling state '{stage}': {e}")
            log_step(f"Login - {stage} Error", "❌ ERROR", str(e))

        time.sleep(2)

    msg = f"Login flow did not complete after {max_stage_transitions} steps | URL: {driver.current_url}"
    print(f"  [LOGIN] FAILED: {msg}")
    log_step("Login - All Attempts Failed", "❌ ERROR", msg)
    return False


def _is_expedia_app_url(url):
    return "vap.expedia.com" in url and not _is_known_login_domain(url)


def _is_sharepoint_app_url(url):
    return "sharepoint.com" in url and "login" not in url.lower() and not _is_known_login_domain(url)


def check_and_login(driver, url) -> bool:
    keyword = url.rstrip('/').split('/')[-1]
    print(f"  [NAV] Navigating to: {keyword}")
    log_step(f"Nav - {keyword[:50]}", "⏳ Wait", f"driver.get({url[:80]}...)")
    driver.get(url)

    try:
        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script("return document.readyState") == "complete")
    except TimeoutException:
        pass
    time.sleep(2)

    print(f"  [NAV] Page loaded | URL: {driver.current_url[:80]}")
    log_step("Nav - Page Loaded", "✅ OK", f"URL after load: {driver.current_url[:120]}")

    try:
        driver.save_screenshot(
            r"C:\Users\huuchinh.nguyen\Concentrix Corporation"
            r"\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\bot_log\screenshot_login.png")
        print(f"  [SCREEN] Screenshot saved")
        log_step("Screenshot", "✅ OK", "screenshot_login.png saved")
    except Exception as ex:
        log_step("Screenshot", "⚠️ WARN", str(ex))

    ok = resolve_login_if_needed(driver, _is_expedia_app_url)
    if not ok:
        raise RuntimeError(f"Login flow did not complete | Current URL: {driver.current_url}")

    if url not in driver.current_url:
        print(f"  [NAV] Redirecting back to target URL...")
        log_step("Nav - Redirect to Target", "⏳ Wait", f"Redirecting to {url[:80]}")
        driver.get(url)
        time.sleep(3)

    try:
        driver.switch_to.alert.accept()
    except NoAlertPresentException:
        pass

    try:
        WebDriverWait(driver, LOGIN_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, LOGIN_VERIFY_CSS)))
        print(f"  [LOGIN] Console page confirmed loaded ({LOGIN_VERIFY_CSS})")
        log_step("Login - Page Confirmed", "✅ OK", f"'{LOGIN_VERIFY_CSS}' found, console ready")
        return True
    except TimeoutException:
        raise RuntimeError(
            f"Console did not load within {LOGIN_TIMEOUT}s\n"
            f"   Current URL: {driver.current_url}"
        )


def verify_on_app(driver, url_keyword, timeout=15):
    print(f"  [VERIFY] Checking page contains: '{url_keyword}'...")
    log_step(f"Verify - {url_keyword[:50]}", "⏳ Wait",
             f"Checking URL contains '{url_keyword}' (timeout={timeout}s)")
    deadline = time.time() + timeout
    while time.time() < deadline:
        cur = driver.current_url
        if (
            "vap.expedia.com" in cur
            and "/login" not in cur
            and "okta.com" not in cur
            and url_keyword in cur
        ):
            print(f"  [VERIFY] OK — confirmed on: {url_keyword}")
            log_step(f"Verify - {url_keyword[:50]}", "✅ OK", f"URL verified: {cur[:120]}")
            return True
        time.sleep(1)
    raise RuntimeError(
        f"Page verification failed: expected '{url_keyword}'\n"
        f"   Current URL: {driver.current_url}"
    )


_sep("INIT CHROME DRIVER")
log_step("Cell 2 Start", "✅ OK", "Cell 2 reached")
t_init = time.time()
driver = None

try:
    print(f"[INIT] Launching Chrome with profile: new_chrome_profile")
    log_step("Init - Launch Chrome", "⏳ Wait", f"Chromedriver: {os.path.basename(CHROMEDRIVER)}")

    chrome_options = Options()
    chrome_options.add_argument(r"--user-data-dir=C:/temp/new_chrome_profile")
    chrome_options.add_argument(r"--profile-directory=Default")
    chrome_options.add_argument("--start-maximized")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--remote-debugging-port=9222")
    chrome_options.add_argument("--safebrowsing-disable-download-protection")
    chrome_options.add_experimental_option("prefs", {
        "safebrowsing.enabled": False,
        "safebrowsing.disable_download_protection": True,
        "profile.default_content_setting_values.automatic_downloads": 1,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "download.default_directory": r"C:\temp\expedia_downloads",
    })

    service = Service(CHROMEDRIVER, service_args=["--log-level=OFF"])
    driver  = webdriver.Chrome(service=service, options=chrome_options)
    driver.get(URL_BREAKDOWN)

    wait    = WebDriverWait(driver, 15)
    wait_sp = WebDriverWait(driver, 20)

    elapsed_init = round(time.time() - t_init, 1)
    print(f"[INIT] Chrome started | Download dir: C:\\temp\\expedia_downloads | {elapsed_init}s")
    log_step("Init Driver", "✅ OK", "Chrome started successfully", elapsed_init)

except Exception as e:
    elapsed = round(time.time() - t_init, 1)
    print(f"[INIT] FAILED: {e}")
    log_step("Init Driver", "❌ ERROR", str(e), elapsed)
    raise

print(f"\n{'═'*60}")
print(f"  BOT RUNNING  |  {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
print(f"{'═'*60}")

try:
    _sep("[1/5] Current Interval CSV")
    t1 = time.time()
    log_step("Step 1 - Start", "⏳ Wait", "Beginning Current Interval download")
    try:
        print(f"  [{_ts()}] check_and_login -> agentBreakdownRealtimeDashboard")
        check_and_login(driver, URL_BREAKDOWN)
        print(f"  [{_ts()}] verify_on_app -> agentBreakdownRealtimeDashboard")
        verify_on_app(driver, "agentBreakdownRealtimeDashboard")
        print(f"  [{_ts()}] Looking for settingsButton on page...")
        log_step("Step 1 - Find Settings Button", "⏳ Wait", "Querying button.settingsButton")
        btns = wait.until(
            lambda d: d.find_elements(By.CSS_SELECTOR, "button.settingsButton"))
        if not btns:
            raise Exception("No settingsButton found on page")
        print(f"  [{_ts()}] Checking for modal overlay before clicking settings...")
        dismiss_modal(driver)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btns[0])
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", btns[0])
        print(f"  [{_ts()}] Settings button clicked — dropdown should open")
        log_step("Step 1 - Click Settings Button", "✅ OK", f"settingsButton clicked at {_ts()}")
        click_download_csv(driver, wait, keyword="Current Interval", step_label="Step 1")
        n = move_files("Current Interval", DIRS["current_interval"], step_label="Step 1")
        elapsed = round(time.time() - t1, 1)
        print(f"  [{_ts()}] STEP 1 COMPLETE | {n} file(s) moved | {elapsed}s")
        log_step("Step 1 - Complete", "✅ OK", f"Current Interval done, {n} file(s) moved", elapsed)
    except Exception as e:
        elapsed = round(time.time() - t1, 1)
        print(f"  [{_ts()}] STEP 1 FAILED ({elapsed}s): {e}")
        log_step("Step 1 - Current Interval", "❌ ERROR", str(e), elapsed)

    _sep("[2/5] Logged-In Agents CSV")
    t2 = time.time()
    log_step("Step 2 - Start", "⏳ Wait", "Beginning Logged-In Agents download")
    try:
        print(f"  [{_ts()}] check_and_login -> agentRealtime")
        check_and_login(driver, URL_REALTIME)
        print(f"  [{_ts()}] verify_on_app -> agentRealtime")
        verify_on_app(driver, "agentRealtime")
        driver.execute_script("document.body.click()"); time.sleep(1)
        print(f"  [{_ts()}] Looking for settingsButton near 'Logged-In Agents'...")
        log_step("Step 2 - Find Settings Button", "⏳ Wait",
                 "JS query for settingsButton near 'Logged-In Agents'")
        btn = wait.until(lambda d: d.execute_script("""
            const el = Array.from(document.querySelectorAll('*')).find(e =>
                e.childNodes.length === 1 &&
                e.childNodes[0].nodeType === Node.TEXT_NODE &&
                e.textContent.trim() === 'Logged-In Agents');
            if (!el) return null;
            let n = el.parentElement;
            while (n && n !== document.body) {
                const b = n.querySelectorAll('button.settingsButton');
                if (b.length === 1) return b[0];
                n = n.parentElement;
            }
            return null;"""))
        if btn is None:
            raise Exception("settingsButton not found near 'Logged-In Agents'")
        print(f"  [{_ts()}] Checking for modal overlay before clicking settings...")
        dismiss_modal(driver)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", btn)
        print(f"  [{_ts()}] Settings button for 'Logged-In Agents' clicked")
        log_step("Step 2 - Click Settings Button", "✅ OK",
                 f"settingsButton near 'Logged-In Agents' clicked at {_ts()}")
        click_download_csv(driver, wait, keyword="Logged-In Agents", step_label="Step 2")
        n = move_files("Logged-In Agents", DIRS["current_agent"], step_label="Step 2")
        elapsed = round(time.time() - t2, 1)
        print(f"  [{_ts()}] STEP 2 COMPLETE | {n} file(s) moved | {elapsed}s")
        log_step("Step 2 - Complete", "✅ OK", f"Logged-In Agents done, {n} file(s) moved", elapsed)
    except Exception as e:
        elapsed = round(time.time() - t2, 1)
        print(f"  [{_ts()}] STEP 2 FAILED ({elapsed}s): {e}")
        log_step("Step 2 - Logged-In Agents", "❌ ERROR", str(e), elapsed)

    _sep("[3/5] Assigned Workitem (Connect) CSV")
    t3 = time.time()
    log_step("Step 3 - Start", "⏳ Wait", "Beginning Assigned Workitem download")
    try:
        print(f"  [{_ts()}] check_and_login -> agentRealtime (guard from Step 2)")
        check_and_login(driver, URL_REALTIME)
        print(f"  [{_ts()}] verify_on_app -> agentRealtime")
        verify_on_app(driver, "agentRealtime")
        driver.execute_script("document.body.click()"); time.sleep(1)
        print(f"  [{_ts()}] Looking for settingsButton near 'Assigned Workitem (Connect)'...")
        log_step("Step 3 - Find Settings Button", "⏳ Wait",
                 "JS query for settingsButton near 'Assigned Workitem (Connect)'")
        btn2 = wait.until(lambda d: d.execute_script("""
            const el = Array.from(document.querySelectorAll('*')).find(e =>
                e.childNodes.length === 1 &&
                e.childNodes[0].nodeType === Node.TEXT_NODE &&
                e.textContent.trim() === 'Assigned Workitem (Connect)');
            if (!el) return null;
            let n = el.parentElement;
            while (n && n !== document.body) {
                const b = n.querySelectorAll('button.settingsButton');
                if (b.length === 1) return b[0];
                n = n.parentElement;
            }
            return null;"""))
        if btn2 is None:
            raise Exception("settingsButton not found near 'Assigned Workitem (Connect)'")
        print(f"  [{_ts()}] Checking for modal overlay before clicking settings...")
        dismiss_modal(driver)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn2)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", btn2)
        print(f"  [{_ts()}] Settings button for 'Assigned Workitem (Connect)' clicked")
        log_step("Step 3 - Click Settings Button", "✅ OK",
                 f"settingsButton near 'Assigned Workitem (Connect)' clicked at {_ts()}")
        click_download_csv(driver, wait, keyword="Assigned Workitem (Connect)", step_label="Step 3")
        n = move_files("Assigned Workitem (Connect)", DIRS["lc_rawdata"], step_label="Step 3")
        elapsed = round(time.time() - t3, 1)
        print(f"  [{_ts()}] STEP 3 COMPLETE | {n} file(s) moved | {elapsed}s")
        log_step("Step 3 - Complete", "✅ OK", f"Assigned Workitem done, {n} file(s) moved", elapsed)
    except Exception as e:
        elapsed = round(time.time() - t3, 1)
        print(f"  [{_ts()}] STEP 3 FAILED ({elapsed}s): {e}")
        log_step("Step 3 - Assigned Workitem", "❌ ERROR", str(e), elapsed)

    _sep("[4/5] SharePoint — EN- UCP.xlsx")
    t4 = time.time()
    log_step("Step 4 - Start", "⏳ Wait", f"Navigating to SharePoint: {URL_SHAREPOINT[:80]}")
    print(f"  [{_ts()}] Navigating to SharePoint URL...")
    driver.get(URL_SHAREPOINT)
    try:
        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script("return document.readyState") == "complete")
    except TimeoutException:
        pass
    time.sleep(3)

    ok4 = resolve_login_if_needed(driver, _is_sharepoint_app_url, max_stage_transitions=14, stage_timeout=25)
    cur_url = driver.current_url
    print(f"  [{_ts()}] SharePoint current URL: {cur_url[:80]}")
    log_step("Step 4 - SharePoint Loaded", "✅ OK" if ok4 else "❌ ERROR",
             f"URL after login-resolve: {cur_url[:120]}")

    if not ok4:
        msg = f"SharePoint still on login page after automatic attempt: {cur_url[:100]}"
        print(f"  [{_ts()}] WARN: {msg}")
        print(f"  [{_ts()}] Action needed: manual check / renew session required")
        log_step("Step 4 - SharePoint Auth", "❌ ERROR", msg, time.time() - t4)
    else:
        try:
            print(f"  [{_ts()}] Searching for file 'EN- UCP' on SharePoint...")
            log_step("Step 4 - Find UCP File", "⏳ Wait", "Locating EN- UCP element on SharePoint page")
            file_el = wait_sp.until(EC.presence_of_element_located((By.XPATH,
                "//span[contains(text(),'EN-') and contains(text(),'UCP')]"
                " | //span[contains(text(),'EN- UCP')]"
                " | //a[contains(@title,'EN-') and contains(@title,'UCP')]")))
            found_name = file_el.text or file_el.get_attribute('title')
            print(f"  [{_ts()}] File found: '{found_name}'")
            log_step("Step 4 - Find UCP File", "✅ OK", f"Element found: '{found_name}'")
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", file_el)
            time.sleep(1)
            print(f"  [{_ts()}] Right-clicking to open context menu...")
            log_step("Step 4 - Right Click UCP", "⏳ Wait", "Dispatching contextmenu event")
            driver.execute_script("""
                arguments[0].dispatchEvent(new MouseEvent('contextmenu', {
                    bubbles: true, cancelable: true,
                    view: window, button: 2, buttons: 2
                }));
            """, file_el)
            time.sleep(2)
            print(f"  [{_ts()}] Context menu opened")
            log_step("Step 4 - Context Menu", "✅ OK", "Right-click context menu triggered")
            print(f"  [{_ts()}] Clicking 'Download' from context menu...")
            dl = wait_sp.until(EC.element_to_be_clickable((By.XPATH,
                "//*[text()='Download' or @aria-label='Download'"
                " or @data-automationid='download']")))
            driver.execute_script("arguments[0].click();", dl)
            print(f"  [{_ts()}] 'Download' clicked | Waiting 12s for file...")
            log_step("Step 4 - Click Download", "✅ OK", f"Download clicked at {_ts()}, waiting 12s")
            time.sleep(12)
            moved = False
            for fp in glob.glob(f"{SOURCE_FOLDER}\\*"):
                if fp.endswith(".crdownload"):
                    continue
                name = os.path.basename(fp).upper()
                if "UCP" in name or ("EN" in name and ".XLSX" in name):
                    if os.path.exists(DST_UCP):
                        os.remove(DST_UCP)
                    shutil.move(fp, DST_UCP)
                    print(f"  [{_ts()}] MOVE: {os.path.basename(fp)} -> {os.path.basename(DST_UCP)}")
                    log_step("Step 4 - Move UCP File", "✅ OK",
                             f"Moved '{os.path.basename(fp)}' -> {os.path.basename(DST_UCP)}")
                    moved = True
            elapsed = round(time.time() - t4, 1)
            if moved:
                print(f"  [{_ts()}] STEP 4 COMPLETE | EN- UCP.xlsx ready | {elapsed}s")
                log_step("Step 4 - Complete", "✅ OK", "EN- UCP.xlsx downloaded and moved", elapsed)
            else:
                msg = "UCP file not found in download folder after 12s"
                print(f"  [{_ts()}] STEP 4 WARN: {msg} | {elapsed}s")
                log_step("Step 4 - Complete", "⚠️ WARN", msg, elapsed)
        except Exception as e:
            elapsed = round(time.time() - t4, 1)
            print(f"  [{_ts()}] STEP 4 FAILED ({elapsed}s): {e}")
            log_step("Step 4 - SharePoint UCP", "❌ ERROR", str(e), elapsed)

    _sep("[5/5] Forecast Realtime CSV (x2)")
    t5 = time.time()
    log_step("Step 5 - Start", "⏳ Wait", "Beginning Forecast Realtime download (2 files)")
    try:
        print(f"  [{_ts()}] check_and_login -> forecastRealtime")
        check_and_login(driver, URL_FORECAST)
        print(f"  [{_ts()}] verify_on_app -> forecastRealtime")
        verify_on_app(driver, "forecastRealtime")

        def _wait_two_settings_btns(d):
            b = d.find_elements(By.CSS_SELECTOR, "button.settingsButton")
            return b if len(b) >= 2 else None

        print(f"  [{_ts()}] Looking for ALL settingsButtons on page...")
        log_step("Step 5 - Find Settings Buttons", "⏳ Wait",
                 "Querying all button.settingsButton on forecastRealtime")
        all_btns = wait.until(_wait_two_settings_btns)
        if not all_btns or len(all_btns) < 2:
            raise Exception(f"Expected >= 2 settingsButtons, found {len(all_btns) if all_btns else 0}")
        print(f"  [{_ts()}] Found {len(all_btns)} settingsButton(s)")
        log_step("Step 5 - Found Settings Buttons", "✅ OK",
                 f"{len(all_btns)} settingsButton(s) found on forecastRealtime")

        print(f"  [{_ts()}] Checking modal before clicking 1st settingsButton...")
        dismiss_modal(driver)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", all_btns[0])
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", all_btns[0])
        print(f"  [{_ts()}] 1st settingsButton clicked")
        log_step("Step 5 - Click Settings Button 1", "✅ OK", f"1st settingsButton clicked at {_ts()}")
        click_download_csv(driver, wait, keyword="In-Progress Interval", step_label="Step 5 File 1")
        n1 = move_files("In-Progress Interval", DIRS["forecast_realtime"], step_label="Step 5 File 1")
        print(f"  [{_ts()}] File 1 done | {n1} file(s) moved -> forecast_realtime/")

        time.sleep(1)
        driver.execute_script("document.body.click()"); time.sleep(1)

        print(f"  [{_ts()}] Re-querying settingsButtons for 2nd click...")
        all_btns2 = wait.until(_wait_two_settings_btns)
        if not all_btns2 or len(all_btns2) < 2:
            raise Exception(f"2nd button query: expected >= 2, found {len(all_btns2) if all_btns2 else 0}")

        print(f"  [{_ts()}] Checking modal before clicking 2nd settingsButton...")
        dismiss_modal(driver)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", all_btns2[1])
        time.sleep(0.5)
        driver.execute_script("arguments[0].click();", all_btns2[1])
        print(f"  [{_ts()}] 2nd settingsButton clicked")
        log_step("Step 5 - Click Settings Button 2", "✅ OK", f"2nd settingsButton clicked at {_ts()}")
        click_download_csv(driver, wait, keyword="Interval Summary", step_label="Step 5 File 2")
        n2 = move_files("Interval Summary", DIRS["forecast_interval_sum"], step_label="Step 5 File 2")
        print(f"  [{_ts()}] File 2 done | {n2} file(s) moved -> forecast_interval_summary/")

        elapsed = round(time.time() - t5, 1)
        print(f"  [{_ts()}] STEP 5 COMPLETE | File1={n1} File2={n2} | {elapsed}s")
        log_step("Step 5 - Complete", "✅ OK",
                 f"Forecast done: In-Progress={n1}, Interval Summary={n2} file(s)", elapsed)
    except Exception as e:
        elapsed = round(time.time() - t5, 1)
        print(f"  [{_ts()}] STEP 5 FAILED ({elapsed}s): {e}")
        log_step("Step 5 - Forecast Realtime", "❌ ERROR", str(e), elapsed)

except RuntimeError as e:
    print(f"\n[FATAL] {e}")
    log_step("FATAL ERROR", "❌ FATAL", str(e))
except WebDriverException as e:
    print(f"\n[WEBDRIVER ERROR] {str(e)[:200]}")
    log_step("WEBDRIVER ERROR", "❌ FATAL", str(e)[:500])
finally:
    if driver is not None:
        try:
            driver.quit()
            print(f"\n[DRIVER] Chrome closed successfully")
            log_step("Driver Quit", "✅ OK", "webdriver.quit() called")
        except Exception as eq:
            print(f"\n[DRIVER] Error closing Chrome: {eq}")
            log_step("Driver Quit", "⚠️ WARN", str(eq))

    finish_time = datetime.now().strftime('%H:%M:%S')
    log_step("BOT FINISH", "✅ OK", f"Bot finished at {finish_time}")

    print(f"\n{'═'*60}")
    print(f"  BOT FINISHED  |  {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
    print(f"  Log saved to  :  {LOG_FILE}")
    print(f"{'═'*60}\n")

# %%
import openpyxl
import polars as pl
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

UCP_FILE = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\EN- UCP.xlsx"
TZ_VNT   = ZoneInfo("Asia/Ho_Chi_Minh")
TZ_PST   = ZoneInfo("America/Los_Angeles")

def fuzzy_site(val):
    s = str(val or '').lower()
    if 'viet' in s: return 'HCM'
    if 'kol'  in s: return 'KOL'
    if 'pun'  in s: return 'PUN'
    if 'cai'  in s: return 'CAI'
    return None

def read_ucp_sheet(wb, sheet_name,
                   req_start=7, req_end=10,
                   extra_start=None, extra_end=None,
                   data_start=3, data_end=50):

    ws = wb[sheet_name]

    req_map = {}
    for c in range(req_start, req_end + 1):
        site = fuzzy_site(ws.cell(row=2, column=c).value)
        if site: req_map[site] = c

    sup_map, mov_map = {}, {}
    if extra_start and extra_end:
        current_group = None
        for c in range(extra_start, extra_end + 1):
            r1 = str(ws.cell(row=1, column=c).value or '').lower()
            if 'support' in r1:                        current_group = 'sup'
            elif 'movement' in r1 or 'cross' in r1:   current_group = 'mov'

            site = fuzzy_site(ws.cell(row=2, column=c).value)
            if site:
                if current_group == 'sup': sup_map[site] = c
                elif current_group == 'mov': mov_map[site] = c

    print(f"  [{sheet_name}] Req:{list(req_map)} | Sup:{list(sup_map)} | Mov:{list(mov_map)}")

    rows = []
    for r in range(data_start, data_end + 1):
        row = {}
        has_val = False
        for site in ['HCM', 'KOL', 'PUN', 'CAI']:
            rv = ws.cell(r, req_map[site]).value    if site in req_map else None
            sv = ws.cell(r, sup_map[site]).value    if site in sup_map else None
            mv = ws.cell(r, mov_map[site]).value    if site in mov_map else None
            row[f'Req_{site}'] = rv
            row[f'Sup_{site}'] = sv
            row[f'Mov_{site}'] = mv
            if rv is not None: has_val = True
        if has_val: rows.append(row)

    return pl.DataFrame(rows) if rows else pl.DataFrame()

def gen_intervals(n_rows):
    today    = datetime.now(TZ_PST).date()
    base_pst = datetime(today.year, today.month, today.day, 0, 0, tzinfo=TZ_PST)
    pst_list, vnt_list = [], []
    for i in range(n_rows):
        pst = base_pst + timedelta(minutes=30*i)
        pst_list.append(pst.strftime("%H:%M"))
        vnt_list.append(pst.astimezone(TZ_VNT).strftime("%H:%M"))
    return pst_list, vnt_list

def attach_ucp_intervals(df, lob):
    pst_list, vnt_list = gen_intervals(len(df))
    return df.with_columns([
        pl.lit(lob).alias("LOB"),
        pl.Series("PST", pst_list),
        pl.Series("VNT", vnt_list),
    ])

wb = openpyxl.load_workbook(UCP_FILE, data_only=True)
print(f"Sheets: {wb.sheetnames}")

df_lg = read_ucp_sheet(wb, "LG Chat", req_start=7, req_end=10, extra_start=21, extra_end=28)
df_lg = attach_ucp_intervals(df_lg, "LG Chat")

df_nl = read_ucp_sheet(wb, "NL Chat", req_start=7, req_end=10, extra_start=22, extra_end=30)
df_nl = attach_ucp_intervals(df_nl, "NL Chat")

df_ucp = pl.concat([df_lg, df_nl], how="diagonal_relaxed").sort(["LOB", "PST"])
print(f"df_ucp shape: {df_ucp.shape}")
print(df_ucp.head(5))

# %%
import pathlib
from datetime import datetime, timedelta

CAPTURE_DIR = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE"

def _mtime_dt(path):
    import os, time
    return datetime(*time.localtime(os.path.getmtime(path))[:6])

def cleanup_old_files(root_folder, keep_days=7):
    cutoff_dt     = datetime.now() - timedelta(days=keep_days)
    total_removed = 0
    total_freed   = 0.0

    root = pathlib.Path(root_folder)
    if not root.exists():
        print(f"[CLEANUP] Root folder not found: {root}")
        return

    subfolders = [f for f in root.iterdir() if f.is_dir()]
    print(f"[CLEANUP] Scanning {len(subfolders)} folder(s) in {root.name} | cutoff={cutoff_dt.strftime('%d-%b-%Y')}")

    for folder in sorted(subfolders):
        removed = 0
        freed   = 0.0
        for fn in folder.glob("*"):
            if fn.name.startswith("_") or fn.suffix.lower() not in (".csv", ".xlsx"):
                continue
            try:
                if _mtime_dt(fn) < cutoff_dt:
                    size_kb = fn.stat().st_size / 1024
                    fn.unlink()
                    freed   += size_kb
                    removed += 1
                    print(f"  [DEL] {folder.name}/{fn.name} ({size_kb:.1f}KB)")
            except Exception as e:
                print(f"  [WARN] Failed to delete {folder.name}/{fn.name}: {e}")

        if removed:
            print(f"  [{folder.name}] removed {removed} file(s), freed {freed:.1f}KB")

        total_removed += removed
        total_freed   += freed

    print(f"\n[CLEANUP] Done | {total_removed} file(s) removed | {total_freed/1024:.2f}MB freed")


cleanup_old_files(CAPTURE_DIR, keep_days=7)

# %%
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options

# CHROMEDRIVER = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\chromedriver-win64\chromedriver.exe"

# chrome_options = Options()
# chrome_options.add_argument(r"--user-data-dir=C:/temp/new_chrome_profile")
# chrome_options.add_argument(r"--profile-directory=Default")
# chrome_options.add_argument("--start-maximized")

# driver = webdriver.Chrome(service=Service(CHROMEDRIVER), options=chrome_options)
# driver.get("https://cnxnice02b.nicecloudsvc.com/wfm/supervisor/reports-generate")


