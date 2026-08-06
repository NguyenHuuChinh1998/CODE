# %%
import os, pathlib, json, requests, time
import pandas as pd
import polars as pl
from datetime import datetime, timedelta
import openpyxl
from zoneinfo import ZoneInfo
import io
import contextlib

UCP_FILE    = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\EN- UCP.xlsx"

SITE_COL_MAP = {
    "Vietnam": "HCM",
    "Kolkata": "KOL",
    "Cairo":   "CAI",
    "Pune":    "PUN",
}

TZ_PST      = ZoneInfo("America/Los_Angeles")

TEAMS_WEBHOOK_URL    = "https://default599e51d62f8c43478e591f795a51a9.8c.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/c24f30c010df45a6a6dac9421643bb34/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=5vWDl18a7-IWSvHuZAWgGtQcwM54nEapSArj4JVPnGg"
DATA_DIR             = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\current_agent"
CURRENT_INTERVAL_DIR = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\current_interval"

HC_PARQUET = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\BI_Task\CODE\Resources\hc_extend_combination.parquet"
IEX_DIR    = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\current_iex"

STATUS_STYLE = {
    "✅ On Schedule": {"bg": "#1B5E20", "fg": "#ffffff"},
    "⚠️ Unscheduled": {"bg": "#E65100", "fg": "#ffffff"},
    "❌ Missed":       {"bg": "#B71C1C", "fg": "#ffffff"},
}

LOB_MAP = {
    "LG Chat": [
        "Chat_OD_EN_Car_Activity","Chat_OD_EN_Lodging",
        "Chat - Global English Lodging Nesting","Chat_Lodging English w Car",
        "Chat_AC_GLB_EN_Lodging_Proficient","Chat_AC_GLB_EN_Car_Activity","Chat_AC_GLB_EN_Lodging_Expert",
    ],
    "NL Chat": [
        "Chat - Global English Non- Lodging Nesting","Chat_OD_EN_Dual_GDS",
        "Chat_AC_GLB_EN_Proficient","Chat_AC_GLB_EN_Expert", "Chat_AC_GLB_EN_NL_Nesting"
    ],
}
LOC_STYLE = {
    "HCM":   {"bg":"#DA251D","fg":"#FFD700"},
    "PUN":   {"bg":"#388E3C","fg":"#ffffff"},
    "KOL":   {"bg":"#1565C0","fg":"#ffffff"},
    "CAI":   {"bg":"#E65100","fg":"#ffffff"},
    "OTHER": {"bg":"#757575","fg":"#ffffff"},
}
LOB_STYLE = {
    "LG Chat": {"bg":"#1565C0","fg":"#ffffff"},
    "NL Chat": {"bg":"#2E7D32","fg":"#ffffff"},
}
NOTE_STYLE = {
    "⚠️ over-break":      {"bg":"#E65100","fg":"#ffffff"},
    "⚠️ over-lunch":      {"bg":"#B71C1C","fg":"#ffffff"},
    "⚠️ available idle":  {"bg":"#F57F17","fg":"#ffffff"},
    "⚠️ offline w/ work": {"bg":"#880E4F","fg":"#ffffff"},
    "⚠️ unproductive":    {"bg":"#37474F","fg":"#ffffff"},
    "🔍 need to check":   {"bg":"#4A148C","fg":"#ffffff"},
    "OK":                 {"bg":"#E8F5E9","fg":"#2E7D32"},
}
CONNECT_STATE_STYLE = {
    "AVAILABLE":    ("#E3F2FD","#1565C0"), "READY":        ("#E3F2FD","#1565C0"),
    "BREAK":        ("#FFE0B2","#BF360C"), "LUNCH":        ("#C8E6C9","#1B5E20"),
    "COACHING":     ("#EDE7F6","#4A148C"), "TRAINING":     ("#E8EAF6","#283593"),
    "TEAM MEETING": ("#E8EAF6","#283593"), "OFFLINEWORK":  ("#FFCCBC","#BF360C"),
    "NOT READY":    ("#FFEBEE","#C62828"), "NOTREADY":     ("#FFEBEE","#C62828"),
    "UNAVAILABLE":  ("#FFEBEE","#C62828"), "ENDOFSHIFT":   ("#FAFAFA","#757575"),
    "LOGIN":        ("#FFF8E1","#F57F17"), "PERSONAL":     ("#FBE9E7","#BF360C"),
}



# %%
def convert_to_datetime(st):
    return datetime(*st[:6])

def load_iex_intervals():
    try:
        hc_ext = (
            pl.read_parquet(HC_PARQUET)
            .select(["Date", "IEX ID", "Employee Name", "LOB", "Email Id", "Supervisor Name"])
            .with_columns([
                pl.col("Date").cast(pl.Date, strict=False),
                pl.col("IEX ID").str.replace_all('"', '').cast(pl.Int64, strict=False)
            ])
            .filter(pl.col("Date") >= pl.date(2026, 1, 1))
        )

        iex_files = []
        for fn in pathlib.Path(IEX_DIR).glob('**/*.*'):
            if fn.name.startswith('_'): continue
            sfx = fn.suffixes
            if not (sfx and sfx[-1].lower() in ['.xlsx', '.csv']): continue
            exp_dt = convert_to_datetime(time.localtime(os.path.getmtime(fn)))
            try:
                if sfx[-1].lower() == '.xlsx':
                    with contextlib.redirect_stderr(io.StringIO()):
                        df = pl.read_excel(fn)
                else:
                    if os.path.getsize(fn) == 0: continue
                    with contextlib.redirect_stderr(io.StringIO()):
                        df = pl.read_csv(fn, infer_schema_length=0)
                if df.is_empty(): continue
                df = df.with_columns([
                    pl.lit(fn.stem).alias('sheet_name'),
                    pl.lit(exp_dt).alias('Export time')
                ])
                iex_files.append(df)
            except Exception as e:
                print(f"  ⚠️ {fn.name}: {e}")

        raw = pl.concat(iex_files, how='diagonal_relaxed') if iex_files else pl.DataFrame()
        if raw.is_empty():
            print("⚠️ No IEX files"); return pl.DataFrame()

        rename_map = {
            "Agent Schedules": "Agent", "__UNNAMED__1": "Date",
            "__UNNAMED__2": "Start_Shift",  "__UNNAMED__3": "End_Shift",
            "__UNNAMED__5": "Scheduled Activity",
            "__UNNAMED__6": "Start_Action",  "__UNNAMED__9": "End_Action"
        }
        IEX = (
            raw.select([c for c in raw.columns if c != "__UNNAMED__4"])
               .rename({k: v for k, v in rename_map.items() if k in raw.columns})
               .with_columns(pl.all().cast(pl.String, strict=False))
        )
        IEX = (
            IEX.with_columns([
                pl.when(pl.col("Agent").str.contains("Generation Date: ", literal=True))
                  .then(pl.col("Agent").str.extract(r"Generation Date: (.+)", 1))
                  .otherwise(None).alias("Generate Date")
            ])
            .with_columns([
                pl.col("Generate Date").fill_null(strategy="backward"),
                pl.col("Agent").fill_null(strategy="forward")
            ])
        )

        IEX_edit = (
            IEX.filter(
                (pl.col("Start_Shift").fill_null("") != "Off") &
                (pl.col("Date").fill_null("") != "Date") &
                ~(pl.col("Date").is_null() & pl.col("Scheduled Activity").is_null())
            )
            .with_columns([
                pl.col("Date").fill_null(strategy="forward"),
                pl.col("Start_Shift").fill_null(strategy="forward"),
                pl.col("End_Shift").fill_null(strategy="forward")
            ])
            .filter(pl.col("Scheduled Activity").is_not_null())
        )

        combined_df = (
            IEX_edit.with_columns([
                pl.col("Export time").cast(pl.String),
                pl.col("Date").cast(pl.String),
                pl.col("Generate Date").cast(pl.String),
                pl.col("Agent").str.extract(r"(\d+)", 1).cast(pl.Int64, strict=False).alias("IEX_ID")
            ])
        )

        max_gd = (
            combined_df.group_by(["IEX_ID", "Date"])
            .agg(pl.col("Generate Date").max().alias("Max_GD"))
        )

        processed = (
            combined_df
            .join(max_gd, on=["IEX_ID", "Date"], how="left")
            .filter(pl.col("Generate Date") == pl.col("Max_GD"))
            .with_columns([
                pl.col("Export time").str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S%.f", strict=False),
                pl.col("Date").str.strip_chars().str.strptime(pl.Date, "%m/%d/%y", strict=False),
                pl.col("Start_Shift").str.strptime(pl.Time, "%I:%M %p", strict=False),
                pl.col("End_Shift").str.strptime(pl.Time,   "%I:%M %p", strict=False),
                pl.col("Start_Action").str.strptime(pl.Time, "%I:%M %p", strict=False),
                pl.col("End_Action").str.strptime(pl.Time,   "%I:%M %p", strict=False),
                pl.col("IEX_ID").cast(pl.Int64, strict=False)
            ])
            .with_columns([
                (pl.col("Start_Shift").dt.strftime("%H%M") + "-" +
                 pl.col("End_Shift").dt.strftime("%H%M")).alias("First Shift")
            ])
            .with_columns([
                pl.col("Start_Action").fill_null(pl.col("Start_Shift")),
                pl.col("End_Action").fill_null(pl.col("End_Shift"))
            ])
            .filter(pl.col("Start_Action").is_not_null() & pl.col("End_Action").is_not_null())
            .with_columns([
                (pl.col("Date").cast(pl.String) + " " + pl.col("Start_Action").cast(pl.String))
                  .str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False).alias("DT_Start"),
                (pl.col("Date").cast(pl.String) + " " + pl.col("End_Action").cast(pl.String))
                  .str.strptime(pl.Datetime, "%Y-%m-%d %H:%M:%S", strict=False).alias("DT_End")
            ])
            .with_columns([
                pl.when(pl.col("DT_End") < pl.col("DT_Start"))
                  .then(pl.col("DT_End") + pl.duration(days=1))
                  .otherwise(pl.col("DT_End")).alias("DT_End")
            ])
            .with_columns([
                pl.col("DT_Start").dt.truncate("30m").alias("IV_Start"),
                pl.when(pl.col("DT_End") == pl.col("DT_End").dt.truncate("30m"))
                  .then(pl.col("DT_End"))
                  .otherwise(pl.col("DT_End").dt.truncate("30m") + pl.duration(minutes=30))
                  .alias("IV_End")
            ])
            .with_columns([
                pl.datetime_ranges(
                    pl.col("IV_Start"), pl.col("IV_End"), interval="30m", closed="left"
                ).alias("VNT_Slots")
            ])
            .explode("VNT_Slots")
            .with_columns([
                pl.col("VNT_Slots")
                  .dt.replace_time_zone("Asia/Ho_Chi_Minh")
                  .dt.convert_time_zone("America/Los_Angeles")
                  .dt.replace_time_zone(None)
                  .alias("PST_Datetime"),
                pl.col("DT_Start").dt.strftime("%H:%M").alias("Start_Action"),
                pl.col("DT_End").dt.strftime("%H:%M").alias("End_Action"),
            ])
            .select([
                "IEX_ID", "Date", "First Shift", "Scheduled Activity",
                "Start_Action", "End_Action", "PST_Datetime"
            ])
            .join(hc_ext, left_on=["Date", "IEX_ID"], right_on=["Date", "IEX ID"], how="left")
            .rename({"Email Id": "Agent Email"}, strict=False)
            .with_columns(pl.col("Agent Email").str.to_lowercase())
        )

        processed = processed.unique(
            subset=["Agent Email", "PST_Datetime", "Scheduled Activity"],
            keep="first"
        )

        print(f"✅ IEX intervals: {len(processed)} rows | "
              f"PST: {processed['PST_Datetime'].min()} → {processed['PST_Datetime'].max()}")
        return processed

    except Exception as e:
        print(f"❌ load_iex_intervals error: {e}")
        import traceback; traceback.print_exc()
        return pl.DataFrame()

def build_schedule_mismatch(outage_snap, intervals_df, pst_now):
    if intervals_df.is_empty():
        return pd.DataFrame(), 0

    sched_now = (
        intervals_df.filter(pl.col("PST_Datetime") == pst_now)
        .unique(subset=["Agent Email", "Scheduled Activity"], keep="first")
    )
    if sched_now.is_empty():
        print(f"⚠️ No schedule data for PST={pst_now}"); return pd.DataFrame(), 0

    hcm_agents = (
        outage_snap
        .filter(pl.col("Location") == "HCM")
        .with_columns([
            pl.col("Agent Email").str.to_lowercase(),
            pl.col("Export time")
              .dt.replace_time_zone("Asia/Ho_Chi_Minh")
              .dt.convert_time_zone("America/Los_Angeles")
              .dt.replace_time_zone(None)
              .dt.truncate("30m")
              .alias("PST_Snap"),
            (pl.col("Export time") -
             pl.duration(seconds=pl.col("Duration (s)").fill_null(0)))
              .dt.strftime("%H:%M")
              .alias("State Since")
        ])
        .filter(pl.col("PST_Snap") == pst_now)
        .select(["Agent Name", "Agent Email", "Connect State", "Duration (s)", "State Since"])
    )
    if hcm_agents.is_empty():
        return pd.DataFrame(), 0

    df_compare = (
        hcm_agents.join(
            sched_now.select([
                "Agent Email", "Employee Name", "Supervisor Name", "LOB",
                "Scheduled Activity", "Start_Action", "End_Action"
            ]),
            on="Agent Email", how="left", suffix="_sched"
        )
        .with_columns(
            pl.col("Agent Name").fill_null(pl.col("Employee Name")).alias("Agent Name")
        )
    )
    match_cond = (
        (
            pl.col("Connect State").is_in(["AVAILABLE", "READY", "OUTBOUNDCALL"]) &
            pl.col("Scheduled Activity").is_in(["Open Time", "Extra Hours"])
        ) |
        ((pl.col("Connect State") == "BREAK") & (pl.col("Scheduled Activity") == "Break")) |
        ((pl.col("Connect State") == "LUNCH") & (pl.col("Scheduled Activity") == "Lunch")) |
        pl.col("Scheduled Activity").is_null()
    )

    mismatches = df_compare.filter(~match_cond).to_pandas()
    if mismatches.empty:
        return pd.DataFrame(), 0

    # ── Helper ────────────────────────────────────────────────
    def parse_hm(s):
        try:
            h, m = map(int, str(s).split(':'))
            return h * 60 + m
        except: return None

    def fmt_min(total_min):
        h, m = divmod(int(total_min) % 1440, 60)
        return f"{h:02d}:{m:02d}"

    def calc_row(row):
        state = str(row['Connect State'])
        sched = str(row['Scheduled Activity']) if pd.notna(row['Scheduled Activity']) else ''

        if state in ['BREAK', 'LUNCH'] and sched in ['Open Time', 'Extra Hours']:
            note = 'Wrong Break' if state == 'BREAK' else 'Wrong Lunch'
        elif state in ['AVAILABLE', 'READY'] and sched == 'Break':
            note = 'Missing Break'
        elif state in ['AVAILABLE', 'READY'] and sched == 'Lunch':
            note = 'Missing Lunch'
        else:
            note = f'Wrong Time ({sched})'

        return pd.Series({'Note': note})

    extra = mismatches.apply(calc_row, axis=1)
    mismatches['Note'] = extra['Note']

    # ── Rename ───────────────────────────────────────────────
    mismatches = mismatches.rename(columns={
        'Supervisor Name':    'Manager',
        'Connect State':      'Current State',
        'Scheduled Activity': 'Scheduled For',
        'Start_Action':       'Sched. Start',
        'End_Action':         'Sched. End',
        'State Since':        'State Since',
        'Connect End':        'Est. End',
    })

    col_order = [
        'LOB', 'Agent Name', 'Manager',
        'Current State', 'Scheduled For',
        'Sched. Start', 'Sched. End',
        'Duration (s)', 'State Since', 'Note'
    ]
    mismatches = (mismatches[[c for c in col_order if c in mismatches.columns]]
                  .sort_values(['LOB', 'Current State', 'Agent Name'])
                  .reset_index(drop=True))

    return mismatches, len(mismatches)

def read_ucp_range(wb, sheet_name, header_row=2, data_start=3, data_end=50):
    ws      = wb[sheet_name]
    headers = [str(ws.cell(row=header_row, column=c).value or f"Col_{c}").strip()
               for c in range(7, 11)]
    rows = []
    for row in ws.iter_rows(min_row=data_start, max_row=data_end, min_col=7, max_col=10):
        rows.append([
            str(cell.value) if cell.value is not None else None
            for cell in row
        ])
    schema = {h: pl.Utf8 for h in headers}
    df = pl.DataFrame(rows, schema=schema, orient="row")
    return df.filter(pl.any_horizontal(pl.all().is_not_null()))

def attach_ucp_intervals(df, lob):
    today    = datetime.now(TZ_PST).date()
    base_pst = datetime(today.year, today.month, today.day, 0, 0, tzinfo=TZ_PST)
    pst_list = [(base_pst + timedelta(minutes=30*i)).strftime("%H:%M") for i in range(len(df))]
    return df.with_columns([pl.Series("PST", pst_list), pl.lit(lob).alias("LOB")])

def get_ucp_req_heads():
    try:
        wb    = openpyxl.load_workbook(UCP_FILE, data_only=True)
        df_lg = attach_ucp_intervals(read_ucp_range(wb, "LG Chat"), "LG Chat")
        df_nl = attach_ucp_intervals(read_ucp_range(wb, "NL Chat"), "NL Chat")
        wb.close()
        df_ucp = pl.concat([df_lg, df_nl], how="diagonal_relaxed")

        now_pst  = datetime.now(TZ_PST)
        curr_pst = f"{now_pst.hour:02d}:{(now_pst.minute // 30) * 30:02d}"
        print(f"✅ UCP loaded | PST interval: {curr_pst}")

        matched = df_ucp.filter(pl.col("PST") == curr_pst)
        if matched.is_empty():
            print(f"⚠️ No UCP row for PST={curr_pst}"); return {}

        site_cols = [c for c in matched.columns if c in SITE_COL_MAP]
        if not site_cols:
            print(f"⚠️ No site columns. Columns: {matched.columns}"); return {}

        result = {}
        for row in matched.iter_rows(named=True):
            lob = row["LOB"]
            for col, loc_code in SITE_COL_MAP.items():
                if col in row and row[col] is not None:
                    try: result[(lob, loc_code)] = int(float(str(row[col]).replace(",","")))
                    except: pass
        return result

    except Exception as e:
        print(f"❌ UCP load error: {e}"); return {}

def input_data(data_dir):
    files = []
    for fn in pathlib.Path(data_dir).glob('**/*.*'):
        if fn.name.startswith('_'): continue
        sfx = fn.suffixes
        if not (sfx and sfx[-1].lower() in ['.xlsx', '.csv']): continue
        exp_dt = convert_to_datetime(time.localtime(os.path.getmtime(fn)))
        try:
            if sfx[-1].lower() == '.xlsx':
                    import warnings
                    with warnings.catch_warnings():
                        warnings.simplefilter("ignore")
                        df = pl.read_excel(fn)
            else:
                if os.path.getsize(fn) == 0: continue
                df = pl.read_csv(fn, infer_schema_length=10000)
            if df.is_empty(): continue
            df = df.with_columns([
                pl.lit(fn.stem).alias('sheet_name'),
                pl.lit(exp_dt).alias('Export time')
            ])
            files.append(df)
        except Exception as e:
            print(f"❌ {fn.name}: {e}")
    return pl.concat(files, how='diagonal_relaxed') if files else pl.DataFrame()

def load_current_interval():
    files=[]
    for fn in pathlib.Path(CURRENT_INTERVAL_DIR).glob("Current Interval*"):
        if fn.suffix.lower() not in ('.csv','.xlsx'): continue
        try:
            exp_dt=convert_to_datetime(time.localtime(os.path.getmtime(fn)))
            df=pl.read_csv(fn,infer_schema_length=10000) if fn.suffix.lower()=='.csv' \
               else pl.read_excel(fn,engine='calamine')
            if df.is_empty(): continue
            df=df.with_columns(pl.lit(exp_dt).alias('Export time'))
            files.append(df)
        except: continue
    if not files: print("⚠️ No Current Interval files found"); return pd.DataFrame()
    combined=pl.concat(files,how='diagonal_relaxed')
    latest=combined['Export time'].max()
    return combined.filter(pl.col('Export time')==latest).drop('Export time').to_pandas()

def str_hms_to_seconds(hms):
    try:
        p=[int(x) for x in str(hms).split(':')]
        if len(p)==3: return p[0]*3600+p[1]*60+p[2]
        if len(p)==2: return p[0]*60+p[1]
        return int(p[0])
    except: return None

def process_outage(df, select_cols):
    df=df.sort(["LOB","Connect State","Duration (s)"],descending=[False,False,True])
    total=df.shape[0]
    avail=[c for c in select_cols if c in df.columns]
    return df.head(50).select(avail).to_pandas(), total

def seconds_to_hms(s):
    try:
        s=int(float(s)); h,r=divmod(s,3600); m,sec=divmod(r,60)
        return f'{h:02d}:{m:02d}:{sec:02d}'
    except: return '—'

def hms_to_seconds(hms):
    try:
        p=str(hms).split(':'); return int(p[0])*3600+int(p[1])*60+int(p[2])
    except: return 0

def get_duration_color(hms_str):
    try:
        s=hms_to_seconds(hms_str)
        if s>3600: return "#b71c1c","#ffffff"
        if s>1800: return "#e53935","#ffffff"
        if s>900:  return "#fb8c00","#ffffff"
        if s>300:  return "#fdd835","#1a1a1a"
    except: pass
    return None, None

def make_bar_cell(s, max_s, color):
    ratio  = s / max_s if max_s > 0 else 0
    filled = int(ratio * 15)
    empty  = 15 - filled
    text_color = color if color else "#43a047"
    bar    = '█' * filled + '░' * empty
    return (
        f'<td style="padding:3px 8px;border:1px solid #ddd;'
        f'font-family:monospace;letter-spacing:1px;'
        f'color:{text_color};font-weight:bold;" nowrap>{bar}</td>'
    )

def build_html_table(df, title, is_global=False, cases=0, summary=''):
    now=datetime.now()-(timedelta(hours=14) if is_global else timedelta(0))
    tz='(PST)' if is_global else '(VNT)'
    sub=f"Updated {now.strftime('%d-%b-%Y')} · {now.strftime('%I:%M %p')} {tz}"
    df=df.copy().drop(columns=[c for c in ['sheet_name','Export time'] if c in df.columns])
    if 'Duration (s)' in df.columns:
        df['Duration']=df['Duration (s)'].apply(seconds_to_hms); df=df.drop(columns=['Duration (s)'])
    cols=list(df.columns)
    dur_idx       = cols.index('Duration')         if 'Duration'         in cols else -1
    lob_idx       = cols.index('LOB')              if 'LOB'              in cols else -1
    note_idx      = cols.index('Note')             if 'Note'             in cols else -1
    state_idx     = (cols.index('Connect State')  if 'Connect State'  in cols else cols.index('Current State')  if 'Current State'  in cols else -1)
    loc_idx       = cols.index('Location')         if 'Location'         in cols else -1
    staff_att_idx = cols.index('Staff Attainment') if 'Staff Attainment' in cols else -1
    agent_def_idx = cols.index('Agent Deficit')    if 'Agent Deficit'    in cols else -1
    req_heads_idx = cols.index('Req Heads [UCP]')  if 'Req Heads [UCP]'  in cols else -1
    heads_def_idx = cols.index('Heads Deficit')    if 'Heads Deficit'    in cols else -1
    status_idx    = cols.index('Status')           if 'Status'           in cols else -1
    sched_for_idx = cols.index('Scheduled For') if 'Scheduled For' in cols else -1

    dur_secs=[hms_to_seconds(str(row[cols[dur_idx]])) for _,row in df.iterrows()] if dur_idx>=0 else [0]*len(df)
    max_s=max(dur_secs) if dur_secs and dur_idx>=0 else 1

    th='bgcolor="#1e3a5f" style="color:#ffffff;padding:7px 12px;border:1px solid #2c4f7c;text-align:left;" nowrap'
    hdrs = []
    for c in cols:
        if '[UCP]' in c:
            base = c.replace('[UCP]', '').strip()
            hdrs.append(f'<th bgcolor="#2E7D32" style="color:#ffffff;padding:7px 12px;border:1px solid #1B5E20;text-align:left;" nowrap>{base} [UCP]</th>')
        else:
            hdrs.append(f'<th {th}>{c}</th>')
    if dur_idx >= 0:
        hdrs.insert(dur_idx + 1, f'<th {th}>Duration Bar</th>')

    def tdp(bg,v): return f'<td bgcolor="{bg}" style="color:#1a1a1a;padding:6px 12px;border:1px solid #ddd;" nowrap>{v}</td>'
    def tdb(bg,fg,v,al='left'): return f'<td bgcolor="{bg}" style="color:{fg};padding:6px 12px;border:1px solid #ddd;font-weight:bold;text-align:{al};" nowrap>{v}</td>'

    rows_html=''
    for i,(_,row) in enumerate(df.iterrows()):
        rbg='#f0f4ff' if i%2==0 else '#ffffff'
        lv=str(row[cols[lob_idx]])  if lob_idx>=0  else ''
        nv=str(row[cols[note_idx]]) if note_idx>=0 else ''
        cells=[]
        for j,col in enumerate(cols):
            v=str(row[col]) if pd.notna(row[col]) else '—'
            if j==dur_idx:
                bg,fg=get_duration_color(v); cells.append(tdb(bg,fg,v) if bg else tdp(rbg,v))
                cells.append(make_bar_cell(dur_secs[i],max_s,bg))
            elif j==loc_idx:
                s=LOC_STYLE.get(v); cells.append(tdb(s['bg'],s['fg'],v,'center') if s else tdp(rbg,v))
            elif j==lob_idx:
                s=LOB_STYLE.get(lv); cells.append(tdb(s['bg'],s['fg'],v,'center') if s else tdp(rbg,v))
            elif j == note_idx:
                color_map = {
                    'Wrong Break':   ('#4A148C', '#ffffff'),
                    'Wrong Lunch':   ('#B71C1C', '#ffffff'),
                    'Missing Break': ('#E65100', '#ffffff'),
                    'Missing Lunch': ('#F57F17', '#1a1a1a'),
                }
                bg, fg = color_map.get(v, (rbg, '#1a1a1a'))
                cells.append(tdb(bg, fg, v, 'center'))
            elif j==state_idx:
                sc=CONNECT_STATE_STYLE.get(v); cells.append(tdb(sc[0],sc[1],v) if sc else tdp(rbg,v))
            elif j==staff_att_idx:
                try:
                    n=float(v.replace('%','').replace(',',''))
                    cells.append(tdb('#1B5E20' if n>=95 else '#B71C1C','#ffffff',v,'center'))
                except: cells.append(tdp(rbg,v))
            elif j==agent_def_idx:
                try:
                    n=float(v.replace(',',''))
                    cells.append(tdb('#1B5E20' if n<=0 else '#B71C1C','#ffffff',v,'center'))
                except: cells.append(tdp(rbg,v))
            elif j==req_heads_idx:
                cells.append(tdb('#1e3a5f','#ffffff',v,'center'))
            elif j==heads_def_idx:
                try:
                    n=float(v.replace(',',''))
                    bg='#1B5E20' if n>=0 else '#B71C1C'
                    display=f"+{int(n)}" if n>=0 else str(int(n))
                    cells.append(tdb(bg,'#ffffff',display,'center'))
                except: cells.append(tdp(rbg,v))
            elif j==status_idx:
                s=STATUS_STYLE.get(v,{'bg':rbg,'fg':'#1a1a1a'}); cells.append(tdb(s['bg'],s['fg'],v,'center'))
            elif j == sched_for_idx:
                sf_color = {
                    'Break':     ('#FFE0B2', '#BF360C'),
                    'Lunch':     ('#C8E6C9', '#1B5E20'),
                    'Open Time': ('#E3F2FD', '#1565C0'),
                    'Training':  ('#EDE7F6', '#4A148C'),
                }
                bg, fg = sf_color.get(v, (rbg, '#1a1a1a'))
                cells.append(tdb(bg, fg, v, 'center'))
            else:
                cells.append(tdp(rbg,v))
        rows_html+=f"<tr>{''.join(cells)}</tr>"

    sh=f'  <span style="font-size:11px;">📊 {summary}</span><br>\n' if summary else ''
    return (f'<p>\n  <b style="color:#c0392b;font-size:16px;">🔴 {title}</b><br>\n'
            f'  <span style="font-size:12px;">{sub} &nbsp;|&nbsp; ⚡ <b>{cases} CASES</b></span><br>\n{sh}</p>\n'
            f'<div style="overflow-x:auto;">\n<table border="1" cellpadding="0" cellspacing="0" '
            f'style="border-collapse:collapse;font-size:12px;font-family:Segoe UI,Arial,sans-serif;">\n'
            f'  <thead><tr>{"".join(hdrs)}</tr></thead>\n  <tbody>{rows_html}</tbody>\n</table>\n</div>')

def send_html_via_webhook(df, title, is_global=False, cases=0, summary='', chunk_size=12):
    if df is None or (hasattr(df, 'empty') and df.empty):
        print(f"⏭️  Skipping '{title}'"); return

    MAX_BYTES = 22 * 1024

    def _post(payload_str, label):
        try:
            r = requests.post(TEAMS_WEBHOOK_URL,
                              headers={'Content-Type': 'application/json'},
                              data=payload_str, timeout=30)
            kb = len(payload_str.encode()) / 1024
            print(f"✅ Sent: '{label}' ({kb:.1f}KB)" if r.status_code in (200, 202)
                  else f"❌ Failed [{r.status_code}]: {r.text[:200]}")
        except Exception as e:
            print(f"❌ {e}")

    def _send_chunk(chunk_df, chunk_title, chunk_cases, chunk_summary):
        html    = build_html_table(chunk_df, chunk_title, is_global, chunk_cases, chunk_summary)
        payload = json.dumps({'html': html})

        if len(payload.encode()) <= MAX_BYTES or len(chunk_df) <= 1:
            _post(payload, chunk_title)
        else:
            mid = max(1, len(chunk_df) // 2)
            _send_chunk(chunk_df.iloc[:mid].reset_index(drop=True),
                        chunk_title + " [A]", chunk_cases, chunk_summary)
            _send_chunk(chunk_df.iloc[mid:].reset_index(drop=True),
                        chunk_title + " [B]", len(chunk_df) - mid, "")

    total    = len(df)
    n_chunks = max(1, (total + chunk_size - 1) // chunk_size)

    for i in range(n_chunks):
        chunk = df.iloc[i * chunk_size:(i + 1) * chunk_size].reset_index(drop=True)
        label = f"{title} ({i+1}/{n_chunks})" if n_chunks > 1 else title
        _send_chunk(chunk, label, len(chunk), summary if i == 0 else "")

# %%
# ── Load data ─────────────────────────────────────────────────
outage_db=input_data(DATA_DIR)
if outage_db.is_empty(): raise RuntimeError("❌ No data loaded")
outage_db=outage_db.sort(["Export time"]).filter(pl.col("Export time")==pl.col("Export time").max())

lob_expr=pl.lit(None).cast(pl.Utf8)
for lbl,qs in LOB_MAP.items():
    lob_expr=pl.when(pl.col("Queue Group / Routing Profile").is_in(qs)).then(pl.lit(lbl)).otherwise(lob_expr)
outage_db=(outage_db
    .with_columns(lob_expr.alias("LOB"))
    .filter(pl.col("LOB").is_in(["NL Chat","LG Chat"]))
    .with_columns(
        pl.when(pl.col("Business Location").str.contains("Ho Chi Minh")).then(pl.lit("HCM"))
        .when(pl.col("Business Location").str.contains("Pune")).then(pl.lit("PUN"))
        .when(pl.col("Business Location").str.contains("Kolkata")).then(pl.lit("KOL"))
        .when(pl.col("Business Location").str.contains("Cairo")).then(pl.lit("CAI"))
        .otherwise(pl.lit("OTHER")).alias("Location"))
    .with_columns(pl.col("Duration").cast(str).map_elements(str_hms_to_seconds,return_dtype=pl.Int64).alias("Duration (s)")))
print(f"✅ {len(outage_db)} rows | {outage_db['Location'].unique().to_list()}")

bl_out=outage_db.filter(pl.col("Connect State").is_in(["BREAK","LUNCH"])).with_columns(
    pl.when((pl.col("Connect State")=="BREAK")&(pl.col("Duration (s)")>900))
        .then(pl.lit("⚠️ over-break"))
    .when((pl.col("Connect State")=="LUNCH")&(pl.col("Location")=="HCM")&(pl.col("Duration (s)")>3600))
        .then(pl.lit("⚠️ over-lunch"))
    .when((pl.col("Connect State")=="LUNCH")&(pl.col("Location")!="HCM")&(pl.col("Duration (s)")>1800))
        .then(pl.lit("⚠️ over-lunch"))
    .otherwise(pl.lit("OK")).alias("Note"))
over_out=bl_out.filter(pl.col("Note")!="OK")

ct_out=outage_db.filter(~pl.col("Connect State").is_in(["BREAK","LUNCH","AVAILABLE","READY","OFFLINEWORK"])).with_columns(
    pl.when(pl.col("Connect State").is_in(["COACHING","TRAINING","TEAM MEETING"]))
    .then(pl.lit("🔍 need to check")).otherwise(pl.lit("⚠️ unproductive")).alias("Note"))

cat_db=outage_db.with_columns(
    pl.when((pl.col("Assigned Workitem Count")>=1)|pl.col("Connect State").is_in(["AVAILABLE","READY"])).then(pl.lit("Available"))
    .when(pl.col("Connect State")=="BREAK").then(pl.lit("Break-Idle"))
    .when(pl.col("Connect State")=="LUNCH").then(pl.lit("Lunch-Idle"))
    .when(pl.col("Connect State").is_in(["COACHING","TEAM MEETING"])).then(pl.lit("Coaching-Idle"))
    .when(pl.col("Connect State")=="TRAINING").then(pl.lit("Training-Idle"))
    .otherwise(pl.lit("Other")).alias("Category"))
pivot_g=cat_db.group_by(["Location","LOB","Category"]).agg(pl.col("Agent Name").n_unique().alias("Count")).pivot(values="Count",index=["Location","LOB"],columns="Category").fill_null(0)
for c in ["Available","Break-Idle","Lunch-Idle","Coaching-Idle","Training-Idle","Other"]:
    if c not in pivot_g.columns: pivot_g=pivot_g.with_columns(pl.lit(0).cast(pl.Int64).alias(c))
pivot_g=pivot_g.select(["Location","LOB","Available","Break-Idle","Lunch-Idle","Coaching-Idle","Training-Idle","Other"]).sort(["LOB","Location"])

def _bl_str(df):
    if df.shape[0]==0: return "No cases"
    cnt=df.group_by(["LOB","Connect State"]).agg(pl.len().alias("Count")).sort(["LOB","Connect State"])
    p={}
    for r in cnt.iter_rows(named=True): p.setdefault(r["LOB"],[]).append(f"{r['Connect State']} ×{r['Count']}")
    return "  |  ".join(f"<b>{k}</b>: {', '.join(v)}" for k,v in sorted(p.items()))

B=["Location","Agent Name","Agent Manager","Connect State","Duration (s)","LOB","Note"]
bl_pd,  bl_n = process_outage(bl_out,  B)
ov_pd,  ov_n = process_outage(over_out,B)
ct_pd,  ct_n = process_outage(ct_out,  B)
pivot_pd = pivot_g.to_pandas()

# ── Merge UCP Req Heads ───────────────────────────────────────
req_heads = get_ucp_req_heads()
if req_heads:
    pivot_pd["Req Heads [UCP]"] = pivot_pd.apply(
    lambda r: req_heads.get((r["LOB"], r["Location"]), 0), axis=1).astype(int)
    pivot_pd["Heads Deficit"] = pivot_pd["Available"] - pivot_pd["Req Heads [UCP]"]
    print(f"✅ UCP merged: {req_heads}")
else:
    print("⚠️ UCP not available — pivot sent without Req Heads")

print(f"IC:{len(pivot_pd)} | BL:{bl_n} | Over:{ov_n} | CT:{ct_n}")

ci_pd=load_current_interval()
send_html_via_webhook(ci_pd,    "Current Interval — Agent Breakdown", cases=len(ci_pd))
send_html_via_webhook(pivot_pd, 'IC Overall — All Sites',             cases=len(pivot_pd))
send_html_via_webhook(bl_pd,    'Lunch / Break',                      cases=bl_n,  summary=_bl_str(bl_out))
send_html_via_webhook(ov_pd,    'Overbreak / Overlunch',              cases=ov_n,  summary=_bl_str(over_out))
send_html_via_webhook(ct_pd,    'Coaching / Training / Unproductive', cases=ct_n)

# ── Schedule Compliance: Connect State vs Scheduled Activity ──
pst_now = (
    outage_db.select(
        pl.col("Export time")
          .dt.replace_time_zone("Asia/Ho_Chi_Minh")
          .dt.convert_time_zone("America/Los_Angeles")
          .dt.replace_time_zone(None)
          .dt.truncate("30m")
          .max()
          .alias("PST")
    ).item()
)
print(f"✅ Current PST interval: {pst_now}")

intervals_df             = load_iex_intervals()
mismatch_pd, mismatch_n  = build_schedule_mismatch(outage_db, intervals_df, pst_now)

if not mismatch_pd.empty:
    working_off = mismatch_pd[
        mismatch_pd["Current State"].isin(["AVAILABLE","READY"]) &
        mismatch_pd["Scheduled For"].isin(["Break","Lunch","Training"])
    ].shape[0]
    resting_on = mismatch_pd[
        mismatch_pd["Current State"].isin(["BREAK","LUNCH"]) &
        ~mismatch_pd["Scheduled For"].isin(["Break","Lunch"])
    ].shape[0]
    mis_summary = f"Working when should rest: {working_off} | Resting when should work: {resting_on}"
else:
    mis_summary = ""

send_html_via_webhook(mismatch_pd, 'Schedule Compliance — HCM',
                      cases=mismatch_n, summary=mis_summary)


