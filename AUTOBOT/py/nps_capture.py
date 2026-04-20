# %%
import os
import io
import time
from datetime import datetime, timedelta
from urllib.parse import urlparse

# Thư viện xử lý ảnh và clipboard
from PIL import Image
import win32clipboard

# Thư viện Selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, NoAlertPresentException, StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager

# %%
# ================= CẤU HÌNH =================
expedia_url = "https://console.vap.expedia.com/analytics-console-user-interface/reports/personal_reports/realtime_chat_nps"
teams_url = "https://teams.microsoft.com/"
group_chat_name = "EXP - Realtime Chat NPS"
report_name = "NPS"
wait_seconds = 180

# Cấu hình Chrome Profile
chrome_options = Options()
chrome_options.add_argument(r'--user-data-dir=C:/temp/new_chrome_profile')
chrome_options.add_argument(r'--profile-directory=Default')
chrome_options.add_argument("--start-maximized")

# --- SỬA: Chỉ dùng 1 trong 2 cách khởi tạo Service ---
# Cách 1: Tự động tải driver (đã ẩn)
# service = Service(ChromeDriverManager().install()) 

# Cách 2: Đường dẫn file exe của bạn (Đang dùng)
chromedriver_path = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\chromedriver-win64\chromedriver.exe"
service = Service(chromedriver_path) 

driver = webdriver.Chrome(service=service, options=chrome_options)
wait = WebDriverWait(driver, 15)

# ================= HÀM HỖ TRỢ =================
if not os.path.exists('screenshots'):
    os.makedirs('screenshots')

def send_to_clipboard(image):
    output = io.BytesIO()
    image.convert("RGB").save(output, "BMP")
    data = output.getvalue()[14:]
    output.close()

    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardData(win32clipboard.CF_DIB, data)
    win32clipboard.CloseClipboard()
    print("Image copied to clipboard!")

def check_and_login(driver, target_url):
    driver.get(target_url)
    time.sleep(5)  # Let the page load

    try:
        # Kiểm tra nút Sign-in nhanh (giảm timeout xuống 5s để đỡ chờ lâu nếu đã login)
        sign_in_button = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="console-okta-sign-in"]'))
        )
        print("🔑 Sign-in required detected! Clicking...")
        sign_in_button.click()
        time.sleep(2)
        
        # Xử lý các bước login tiếp theo
        try:
            keep_signed_in_label = WebDriverWait(driver, 5).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="input36"][data-se-for-name="rememberMe"]'))
            )
            keep_signed_in_label.click()
        except TimeoutException:
            pass # Không thấy thì bỏ qua

        try:
            next_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'input.button.button-primary[type="submit"][value="Next"]'))
            )
            next_button.click()
            time.sleep(8)
        except TimeoutException:
            print("No 'Next' button found. Skipping.")
            
        try:
            alert = driver.switch_to.alert
            alert.accept()
        except NoAlertPresentException:
            pass
            
        print("🎉 Login successful! Reloading...")
        driver.get(target_url)
    except TimeoutException:
        print("✅ No sign-in required / Already logged in.")

# ================= BẮT ĐẦU QUY TRÌNH =================

# 1. Mở Expedia và Login
check_and_login(driver, expedia_url)

# 2. Chờ báo cáo load xong
parsed_url = urlparse(expedia_url)
path_fragment = parsed_url.path.split('/')[-1]
print(f"Waiting for URL to contain: {path_fragment}")
try:
    wait.until(EC.url_contains(path_fragment))
    print(f"URL matched. Waiting {wait_seconds} seconds for data rendering...")
    time.sleep(wait_seconds)
except TimeoutException:
    print("Warning: URL did not match expected fragment. Proceeding anyway...")

# 3. Chụp màn hình
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
screenshot_path = f'screenshots/screenshot_{timestamp}.png'

driver.save_screenshot(screenshot_path)
print(f"Screenshot saved at: {screenshot_path}")

image = Image.open(screenshot_path)
send_to_clipboard(image)

# 4. Mở Teams và gửi tin nhắn
print("Navigate to Teams...")
driver.get(teams_url)
time.sleep(5) # Chờ Teams load khung cơ bản

# --- SỬA: Logic tìm Group Chat (Retry + XPath Title) ---
print(f"Finding group: {group_chat_name}")

# XPath này tìm thẻ span có thuộc tính title chính xác (ổn định hơn text)
# Hoặc tìm text chính xác
search_xpath = f"//span[@title='{group_chat_name}'] | //span[text()='{group_chat_name}']"

found_group = False
for i in range(3): # Thử lại 3 lần
    try:
        # Chờ element xuất hiện và có thể click
        element = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.XPATH, search_xpath))
        )
        
        # Nếu có nhiều element trùng tên, ưu tiên cái nằm ở Sidebar (tọa độ X nhỏ)
        all_elements = driver.find_elements(By.XPATH, search_xpath)
        target_elem = None
        
        for e in all_elements:
            if e.is_displayed() and e.location['x'] < 500: # Sidebar thường ở bên trái (< 500px)
                target_elem = e
                break
        
        if target_elem:
            target_elem.click()
        else:
            # Nếu không lọc được theo tọa độ, click cái đầu tiên tìm thấy
            element.click()
            
        print("✅ Clicked group chat successfully.")
        found_group = True
        break
    except StaleElementReferenceException:
        print(f"⚠️ Stale Element (Lần {i+1}). Teams đang reload DOM. Chờ 3s...")
        time.sleep(3)
    except TimeoutException:
        print("❌ Timeout: Không tìm thấy group chat.")
        break
    except Exception as e:
        print(f"❌ Lỗi khác: {e}")
        break

if found_group:
    time.sleep(3) # Chờ khung chat load sau khi click

    # Tìm khung nhập liệu
    try:
        chat_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div[role='textbox']")))
        chat_box.click()
        
        # Soạn tin nhắn
        now_time = datetime.now()
        realtime = now_time.strftime("%I:%M %p")
        chat_text = f"The {report_name} updated at {realtime} (VNT)"

        actions = ActionChains(driver)
        actions.send_keys(chat_text)
        actions.perform()
        
        time.sleep(1)

        # Paste ảnh (Ctrl+V)
        actions = ActionChains(driver)
        actions.key_down(Keys.CONTROL).send_keys('v').key_up(Keys.CONTROL).perform()
        
        # Chờ ảnh upload lên Teams (quan trọng)
        print("Waiting for image upload...")
        time.sleep(5) 

        # Gửi (Enter)
        actions.send_keys(Keys.ENTER).perform()
        print("Message sent!")
        
        time.sleep(5)
    except Exception as e:
        print(f"Lỗi khi gửi tin nhắn: {e}")

print("Done. Closing driver.")
driver.quit()

# %%
from openpyxl import Workbook, load_workbook
from datetime import datetime

end_time = datetime.now()
execution_time = (end_time - start_time).total_seconds()

log_file = "bot_log/bot_capture_log.xlsx"

if not os.path.exists(log_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["BOT Name", "Run at", "End at", "Execution Time (seconds)"])
else:
    wb = load_workbook(log_file)
    ws = wb["Log"]

ws.append([
    report_name,
    start_time.strftime('%Y-%m-%d %H:%M:%S'),
    end_time.strftime('%Y-%m-%d %H:%M:%S'),
    execution_time
])

wb.save(log_file)


