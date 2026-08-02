# %%
import os, time, shutil, glob
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, NoAlertPresentException, WebDriverException
from datetime import datetime, timedelta

# ── CONFIG ─────────────────────────────────────────────────────
CHROMEDRIVER   = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\chromedriver-win64\chromedriver.exe"
SOURCE_FOLDER  = r"C:\temp\expedia_downloads"
BASE_CAPTURE   = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE"
DIRS = {
    "current_agent"   : os.path.join(BASE_CAPTURE, "current_agent"),
    "lc_rawdata"      : os.path.join(BASE_CAPTURE, "lc_rawdata_in_console"),
    "current_interval": os.path.join(BASE_CAPTURE, "current_interval"),
}
for d in DIRS.values(): os.makedirs(d, exist_ok=True)

URL_BREAKDOWN  = "https://console.vap.expedia.com/analytics-console-user-interface/optics/agentBreakdownRealtimeDashboard"
URL_REALTIME   = "https://console.vap.expedia.com/analytics-console-user-interface/optics/agentRealtime"
URL_SHAREPOINT = (
    "https://cnxmail-my.sharepoint.com/shared?listurl=https%3A%2F%2Fcnxmail-my%2E"
    "sharepoint%2Ecom%2Fpersonal%2Fahmed_ahmedkamh_concentrix_com%2FDocuments"
    "&id=%2Fpersonal%2Fahmed_ahmedkamh_concentrix_com%2FDocuments"
)
DST_UCP          = os.path.join(BASE_CAPTURE, "EN- UCP.xlsx")
LOGIN_VERIFY_CSS = "button.settingsButton"
LOGIN_TIMEOUT    = 20
CNX_USER         = "huuchinh.nguyen@concentrix.com"
CNX_PASS         = "Vuthihongtham@130499"

# ── HELPERS ────────────────────────────────────────────────────
def move_files(keyword, dest_dir):
    moved = 0
    for pat in [f"{SOURCE_FOLDER}\\{keyword}*.csv", f"{SOURCE_FOLDER}\\{keyword}*.xlsx"]:
        for fp in glob.glob(pat):
            if fp.endswith(".crdownload"): continue
            dst = os.path.join(dest_dir, os.path.basename(fp))
            if os.path.exists(dst): os.remove(dst)
            shutil.move(fp, dst)
            print(f"  📁 Moved: {os.path.basename(fp)}"); moved += 1
    if not moved: print(f"  ⚠️ No file '{keyword}*' found")
    return moved

def click_download_csv(driver, wait, keyword=None, timeout=30):
    wait.until(EC.presence_of_element_located(
        (By.CSS_SELECTOR, "div.uitk-menu-container[aria-hidden='false']")))
    wait.until(EC.element_to_be_clickable((By.XPATH,
        "//div[contains(@class,'uitk-menu-open')][@aria-hidden='false']"
        "//span[text()='Download CSV']/ancestor::button"))).click()
    print("  ✅ Clicked Download CSV")
    if keyword:
        start = time.time()
        while time.time() - start < timeout:
            matches = [f for f in
                glob.glob(f"{SOURCE_FOLDER}\\{keyword}*.csv") +
                glob.glob(f"{SOURCE_FOLDER}\\{keyword}*.xlsx")
                if not f.endswith('.crdownload')]
            if matches:
                time.sleep(0.5)
                print(f"  ⚡ File ready in {round(time.time()-start,1)}s")
                return
            time.sleep(0.5)
        print(f"  ⚠️ Timeout {timeout}s")
    else:
        time.sleep(8)

def _fill_input(driver, element, value):
    """Clear hoàn toàn rồi điền giá trị mới — an toàn với MUI input."""
    driver.execute_script("arguments[0].value = '';", element)
    element.click()
    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.DELETE)
    time.sleep(0.3)
    element.send_keys(value)

def cnx_auth_login(driver):
    """
    Xử lý luồng Concentrix Authentication:
    Concentrix Authentication → USERNAME tab → email → Next
    → Password card → password → Next → chờ redirect
    """
    try:
        # Step 0: Click "Concentrix Authentication"
        try:
            cnx_btn = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable((By.XPATH,
                    '//span[contains(@class,"largeTextNoWrap") and '
                    'contains(text(),"Concentrix Authentication")]'
                ))
            )
            driver.execute_script("arguments[0].click();", cnx_btn)
            print("  ✅ Clicked 'Concentrix Authentication'")
            time.sleep(3)
        except TimeoutException:
            pass

        # Step 1: Click USERNAME tab
        username_tab = WebDriverWait(driver, 8).until(
            EC.element_to_be_clickable((By.XPATH,
                '//button[@aria-label="Passwordless users, login here." '
                'and normalize-space(text())="Username"]'
            ))
        )
        driver.execute_script("arguments[0].click();", username_tab)
        print("  ✅ Clicked 'Username' tab")
        time.sleep(1)

        # Step 2: Remember me checkbox
        try:
            cb = driver.find_element(By.ID, "checkboxRememberMe")
            if not cb.is_selected():
                driver.execute_script("arguments[0].click();", cb)
                print("  ✅ Checked 'Remember me on this device'")
            time.sleep(0.5)
        except Exception:
            pass

        # Step 3: username
        user_input = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "username"))
        )
        _fill_input(driver, user_input, CNX_USER)
        print(f"  ✅ Entered username: {CNX_USER}")
        time.sleep(0.5)

        # Step 4: Click Next (username)
        next_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[type="submit"]'))
        )
        driver.execute_script("arguments[0].click();", next_btn)
        print("  ✅ Clicked Next (username)")
        time.sleep(2)

        # Step 5: Click "Password" method card
        try:
            pwd_card = WebDriverWait(driver, 8).until(
                EC.element_to_be_clickable((By.XPATH,
                    '//*[contains(@class,"jss142") and normalize-space(text())="Password"]'
                    '/ancestor::div[contains(@class,"jss140")]'
                ))
            )
            driver.execute_script("arguments[0].click();", pwd_card)
            print("  ✅ Clicked 'Password' method card")
            time.sleep(2)
        except TimeoutException:
            pass

        # Step 6: password
        pwd_input = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.ID, "password"))
        )
        _fill_input(driver, pwd_input, CNX_PASS)
        print("  ✅ Entered password")
        time.sleep(0.5)

        # Step 7: Click Next (password)
        next_btn2 = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[type="submit"]'))
        )
        driver.execute_script("arguments[0].click();", next_btn2)
        print("  ✅ Clicked Next (password)")
        time.sleep(3)

        print("  🎉 Concentrix Authentication completed")

    except TimeoutException:
        print("  ✅ No Concentrix Auth prompt — skipping")

def check_and_login(driver, url, manual_login_timeout=120) -> bool:
    print(f"  🌐 Navigating to: {url.split('/')[-1]}")
    driver.get(url); time.sleep(10)

    # Case 1: Okta sign-in button (Expedia SSO)
    try:
        sign_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'button[data-testid="console-okta-sign-in"]')))
        print("  🔑 Okta sign-in detected, clicking...")
        sign_btn.click(); time.sleep(2)

        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'label[for="input36"][data-se-for-name="rememberMe"]'))).click()
            time.sleep(1)
        except TimeoutException: pass

        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'input.button.button-primary[type="submit"][value="Next"]'))).click()
            print("  ✅ Clicked Next (Okta username)")
        except TimeoutException: pass

        deadline = time.time() + manual_login_timeout
        while time.time() < deadline:
            if "expedia.com" in driver.current_url or "vap.expedia" in driver.current_url:
                print("  🎉 Okta login completed")
                break
            time.sleep(2)
        else:
            print(f"  ⚠️ Okta timeout | URL: {driver.current_url}")

        try: driver.switch_to.alert.accept()
        except NoAlertPresentException: pass

        if url not in driver.current_url:
            driver.get(url); time.sleep(5)

    except TimeoutException:
        # Case 2: Concentrix Authentication page
        try:
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH,
                    '//span[contains(@class,"largeTextNoWrap") and '
                    'contains(text(),"Concentrix Authentication")]'
                ))
            )
            print("  🔑 Concentrix Auth page detected")
            cnx_auth_login(driver)

            deadline = time.time() + manual_login_timeout
            while time.time() < deadline:
                if "expedia.com" in driver.current_url or "vap.expedia" in driver.current_url:
                    print("  🎉 Login completed — redirected to Expedia")
                    break
                time.sleep(2)
            else:
                print(f"  ⚠️ Login timeout | URL: {driver.current_url}")

            if url not in driver.current_url:
                driver.get(url); time.sleep(5)

        except TimeoutException:
            print("  ✅ No login prompt — already authenticated")

    # Verify console loaded
    try:
        WebDriverWait(driver, LOGIN_TIMEOUT).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, LOGIN_VERIFY_CSS)))
        print("  ✅ Console page confirmed loaded")
        return True
    except TimeoutException:
        raise RuntimeError(
            f"❌ Console did NOT load within {LOGIN_TIMEOUT}s\n"
            f"   Current URL: {driver.current_url}"
        )

# ── INIT DRIVER ────────────────────────────────────────────────
chrome_options = Options()
chrome_options.add_argument(r"--user-data-dir=C:/temp/new_chrome_profile")
chrome_options.add_argument(r"--profile-directory=Default")
chrome_options.add_argument("--start-maximized")
driver  = webdriver.Chrome(service=Service(CHROMEDRIVER), options=chrome_options)
wait    = WebDriverWait(driver, 15)
wait_sp = WebDriverWait(driver, 20)

print(f"\n{'═'*55}")
print(f"🚀 Bot started: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
print(f"{'═'*55}")

try:
    # ══ STEP 1: Current Interval ═══════════════════════════════
    print("\n[1/4] Current Interval CSV")
    check_and_login(driver, URL_BREAKDOWN)
    try:
        btns = wait.until(lambda d: d.find_elements(By.CSS_SELECTOR, "button.settingsButton"))
        if not btns: raise Exception("No settingsButton found")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btns[0]); time.sleep(0.5)
        driver.execute_script("arguments[0].click();", btns[0])
        click_download_csv(driver, wait, keyword="Current Interval")
        move_files("Current Interval", DIRS["current_interval"])
    except Exception as e:
        print(f"  ❌ Step 1 failed: {e}")

    # ══ STEP 2: Logged-In Agents ═══════════════════════════════
    print("\n[2/4] Logged-In Agents CSV")
    check_and_login(driver, URL_REALTIME)
    try:
        # Dismiss any open menu first
        driver.execute_script("document.body.click();")
        time.sleep(1)

        btn = wait.until(lambda d: d.execute_script("""
            const el=Array.from(document.querySelectorAll('*')).find(e=>
                e.childNodes.length===1&&e.childNodes[0].nodeType===Node.TEXT_NODE&&
                e.textContent.trim()==='Logged-In Agents');
            if(!el)return null;
            let n=el.parentElement;
            while(n&&n!==document.body){
                const b=n.querySelectorAll('button.settingsButton');
                if(b.length===1)return b[0]; n=n.parentElement;}
            return null;"""))
        if btn is None: raise Exception("settingsButton not found")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", btn)
        click_download_csv(driver, wait, keyword="Logged-In Agents")
        move_files("Logged-In Agents", DIRS["current_agent"])
    except Exception as e:
        print(f"  ❌ Step 2 failed: {e}")

    # ══ STEP 3: Assigned Workitem (Connect) ════════════════════
    print("\n[3/4] Assigned Workitem (Connect) CSV")
    try:
        # Dismiss any open menu first
        driver.execute_script("document.body.click();")
        time.sleep(1)

        btn2 = wait.until(lambda d: d.execute_script("""
            const el=Array.from(document.querySelectorAll('*')).find(e=>
                e.childNodes.length===1&&e.childNodes[0].nodeType===Node.TEXT_NODE&&
                e.textContent.trim()==='Assigned Workitem (Connect)');
            if(!el)return null;
            let n=el.parentElement;
            while(n&&n!==document.body){
                const b=n.querySelectorAll('button.settingsButton');
                if(b.length===1)return b[0]; n=n.parentElement;}
            return null;"""))
        if btn2 is None: raise Exception("settingsButton not found")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn2)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", btn2)
        click_download_csv(driver, wait, keyword="Assigned Workitem (Connect)")
        move_files("Assigned Workitem (Connect)", DIRS["lc_rawdata"])
    except Exception as e:
        print(f"  ❌ Step 3 failed: {e}")

    # ══ STEP 4: SharePoint — EN- UCP.xlsx ══════════════════════
    print("\n[4/4] SharePoint — EN- UCP.xlsx")
    driver.get(URL_SHAREPOINT); time.sleep(10)
    try:
        file_el = wait_sp.until(EC.presence_of_element_located((By.XPATH,
            "//span[contains(text(),'EN-') and contains(text(),'UCP')]"
            " | //span[contains(text(),'EN- UCP')]"
            " | //a[contains(@title,'EN-') and contains(@title,'UCP')]")))
        print(f"  ✅ Found: {file_el.text or file_el.get_attribute('title')}")
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", file_el); time.sleep(1)
        driver.execute_script("""
            arguments[0].dispatchEvent(new MouseEvent('contextmenu',{
                bubbles:true,cancelable:true,view:window,button:2,buttons:2}));
        """, file_el); time.sleep(2)
        dl = wait_sp.until(EC.element_to_be_clickable((By.XPATH,
            "//*[text()='Download' or @aria-label='Download' or @data-automationid='download']")))
        driver.execute_script("arguments[0].click();", dl)
        print("  ✅ Clicked Download"); time.sleep(12)
        moved = False
        for fp in glob.glob(f"{SOURCE_FOLDER}\\*"):
            if fp.endswith(".crdownload"): continue
            name = os.path.basename(fp).upper()
            if "UCP" in name or ("EN" in name and ".XLSX" in name):
                if os.path.exists(DST_UCP): os.remove(DST_UCP)
                shutil.move(fp, DST_UCP)
                print(f"  📁 Moved → {os.path.basename(DST_UCP)}"); moved = True
        if not moved: print("  ⚠️ UCP file not found")
    except Exception as e:
        print(f"  ❌ Step 4 failed: {e}")

except RuntimeError as e:
    print(f"\n🚨 FATAL: {e}")
except WebDriverException as e:
    print(f"\n🚨 WEBDRIVER ERROR: {e}")
finally:
    driver.quit()
    print(f"\n{'═'*55}")
    print(f"✅ Bot finished: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}")
    print(f"{'═'*55}")

# %%
import openpyxl
import polars as pl
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

UCP_FILE = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\EN- UCP.xlsx"
TZ_VNT   = ZoneInfo("Asia/Ho_Chi_Minh")
TZ_PST   = ZoneInfo("America/Los_Angeles")

def read_range(wb, sheet_name, header_row=2, data_start=3, data_end=50):
    ws      = wb[sheet_name]
    headers = [str(ws.cell(row=header_row, column=c).value or f"Col_{c}").strip()
               for c in range(7, 11)]
    rows = []
    for row in ws.iter_rows(min_row=data_start, max_row=data_end, min_col=7, max_col=10):
        rows.append([cell.value for cell in row])
    df = pl.DataFrame(rows, schema=headers, orient="row")
    return df.filter(pl.any_horizontal(pl.all().is_not_null()))

def gen_intervals(n_rows):
    today    = datetime.now(TZ_PST).date()
    base_pst = datetime(today.year, today.month, today.day, 0, 0, tzinfo=TZ_PST)
    vnt_list, pst_list = [], []
    for i in range(n_rows):
        pst = base_pst + timedelta(minutes=30*i)
        vnt = pst.astimezone(TZ_VNT)
        pst_list.append(pst.strftime("%H:%M"))
        vnt_list.append(vnt.strftime("%H:%M"))
    return vnt_list, pst_list

def attach_intervals(df, lob):
    vnt_list, pst_list = gen_intervals(len(df))
    return df.with_columns([
        pl.Series("VNT", vnt_list),
        pl.Series("PST", pst_list),
        pl.lit(lob).alias("LOB"),
    ]).select(["LOB","VNT","PST"] + df.columns)

wb = openpyxl.load_workbook(UCP_FILE, data_only=True)
print(f"Sheets: {wb.sheetnames}")

df_nl = attach_intervals(read_range(wb, "NL Chat"), "NL Chat")
df_lg = attach_intervals(read_range(wb, "LG Chat"), "LG Chat")

df_ucp = pl.concat([df_lg, df_nl], how="diagonal_relaxed").sort(["LOB","PST"])
print(f"df_ucp: {df_ucp.shape}")
print(df_ucp)

# %%
# from selenium import webdriver
# from selenium.webdriver.chrome.service import Service
# from selenium.webdriver.chrome.options import Options

# CHROMEDRIVER = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\chromedriver-win64\chromedriver.exe"

# chrome_options = Options()
# chrome_options.add_argument(r"--user-data-dir=C:/temp/new_chrome_profile")
# chrome_options.add_argument(r"--profile-directory=Default")
# chrome_options.add_argument("--start-maximized")

# driver = webdriver.Chrome(service=Service(CHROMEDRIVER), options=chrome_options)
# driver.get("https://cnxnice02b.nicecloudsvc.com/wfm/supervisor/reports-generate")


