# %%
import os, time, pathlib, json, requests
import polars as pl
import pandas as pd
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
import openpyxl
import warnings

# [REAL — bỏ comment khi test OK, xóa block DRAFT bên dưới]
# TEAMS_WEBHOOK_URL = "https://default599e51d62f8c43478e591f795a51a9.8c.environment.api.powerplatform.com:443/powerautomate/automations/direct/workflows/d9dfae822f4941d0be070dd295d55658/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=zQ76qlawVl-CtgQ1Okym9_Vz4rdbSAa0Mc7VHESH3N4"

# [DRAFT — đang active để test, xóa block này khi test OK]
TEAMS_WEBHOOK_URL = (
    "https://default599e51d62f8c43478e591f795a51a9.8c.environment.api.powerplatform.com:443"
    "/powerautomate/automations/direct/workflows/30e46f2733bc4e48a92dc32f90ba9329"
    "/triggers/manual/paths/invoke?api-version=1&sp=%2Ftriggers%2Fmanual%2Frun"
    "&sv=1.0&sig=xkD_H8_VvQh_XzhybXDxV3_gWFyC0E4-3Bpe_MJDJ44"
)

FORECAST_DIR          = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\forecast_realtime"
INTERVAL_DIR          = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\current_interval"
UCP_FILE              = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\EN- UCP.xlsx"
FORECAST_INTERVAL_DIR = r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\forecast_interval_summary"

TZ_VNT = ZoneInfo("Asia/Ho_Chi_Minh")
TZ_PST = ZoneInfo("America/Los_Angeles")

CHILD_LOB_MAP = {
    "GEN_GEN_EN_GCS_GLG_CHT_Concentrix (Ho Chi Minh City)": "LG Chat HCM",
    "GEN_GEN_EN_GCS_GLG_CHT_Concentrix (Kolkata)":          "LG Chat Kolkata",
    "GEN_GEN_EN_GCS_GLG_CHT_Concentrix (Pune)":             "LG Chat Pune",
    "GEN_GEN_EN_GCS_GLG_CHT_Concentrix (Cairo)":            "LG Chat Cairo",
}
SITES             = ["HCM", "Kolkata", "Pune", "Cairo"]
INTERVAL_LG_KEY   = "GEN_GEN_EN_GCS_GLG_CHT"
UCP_SITE_COL_MAP  = {"Vietnam": "HCM", "Kolkata": "Kolkata", "Pune": "Pune", "Cairo": "Cairo"}
IC_FAIL_THRESHOLD = 0.95

LOC_STYLE = {
    "HCM":     {"bg": "#DA251D", "fg": "#FFD700"},
    "Kolkata": {"bg": "#1565C0", "fg": "#ffffff"},
    "Pune":    {"bg": "#388E3C", "fg": "#ffffff"},
    "Cairo":   {"bg": "#F57F17", "fg": "#ffffff"},
}
SITE_HEADER_BG = {
    "HCM":     "#B71C1C",
    "Kolkata": "#0D47A1",
    "Pune":    "#1B5E20",
    "Cairo":   "#E65100",
}

# %%
def _mtime_dt(path):
    return datetime(*time.localtime(os.path.getmtime(path))[:6])


def _parse_time_str(t_str, base_date, now_vnt):
    try:
        parts = [int(x) for x in str(t_str).strip().split(":")]
        h  = parts[0] % 24
        m  = parts[1] if len(parts) > 1 else 0
        dt = datetime(base_date.year, base_date.month, base_date.day, h, m, 0)
        if dt > now_vnt + timedelta(hours=12):
            dt -= timedelta(days=1)
        return dt
    except:
        return None


def _parse_ic_time(t_str):
    """Parse Interval Time from forecast_interval_summary (e.g. '8/19/2026 1:30')."""
    if t_str is None:
        return None
    s = str(t_str).strip()
    for fmt in ("%m/%d/%Y %H:%M", "%m/%d/%Y %I:%M %p",
                "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def load_latest_snapshot(folder, glob_pattern="*"):
    files = []
    for fn in pathlib.Path(folder).glob(glob_pattern):
        if fn.name.startswith("_"):
            continue
        if fn.suffix.lower() not in (".csv", ".xlsx"):
            continue
        try:
            exp_dt = _mtime_dt(fn)
            if fn.suffix.lower() == ".xlsx":
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore")
                    df = pl.read_excel(fn)
            else:
                if os.path.getsize(fn) == 0:
                    continue
                df = pl.read_csv(fn, infer_schema_length=10000)
            if df.is_empty():
                continue
            df = df.with_columns(pl.lit(exp_dt).alias("_exp_dt"))
            files.append(df)
        except Exception as e:
            print(f"  [WARN] {fn.name}: {e}")
    if not files:
        return pl.DataFrame(), None
    combined = pl.concat(files, how="diagonal_relaxed")
    latest   = combined["_exp_dt"].max()
    result   = combined.filter(pl.col("_exp_dt") == latest).drop("_exp_dt")
    print(f"  [LOAD] {folder.split(chr(92))[-1]} | {len(result)} rows | export={latest}")
    return result, latest


def load_all_ic_csv(folder):
    path      = pathlib.Path(folder)
    now_pst   = datetime.now(TZ_PST).replace(tzinfo=None)
    cutoff_dt = datetime.combine(now_pst.date() - timedelta(days=1), datetime.min.time())

    dfs = []
    for fn in sorted(path.glob("*.csv")):
        if fn.name.startswith("_"):
            continue
        try:
            if os.path.getsize(fn) == 0:
                continue
            if _mtime_dt(fn) < cutoff_dt:
                continue
            df = pl.read_csv(fn, infer_schema_length=10000)
            if not df.is_empty():
                dfs.append(df)
        except Exception as e:
            print(f"  [WARN-IC] {fn.name}: {e}")

    if not dfs:
        print(f"  [IC] No CSV found in range: {path.name}")
        return pl.DataFrame()

    combined = pl.concat(dfs, how="diagonal_relaxed")
    key_cols = [c for c in ["Interval Time", "Child Forecast Group"] if c in combined.columns]
    if key_cols:
        combined = combined.unique(subset=key_cols, keep="last")
    print(f"  [IC] {len(combined)} unique rows from {len(dfs)} file(s)")
    return combined


def read_ucp_req(ucp_path):
    wb = openpyxl.load_workbook(ucp_path, data_only=True)
    ws = wb["LG Chat"]
    site_col_idx = {}
    for c in range(7, 12):
        hdr = str(ws.cell(row=2, column=c).value or "").strip()
        for k, v in UCP_SITE_COL_MAP.items():
            if k.lower() in hdr.lower():
                site_col_idx[c] = v
                break
    rows = []
    for i, r in enumerate(range(3, 51)):
        pst_t = f"{(i // 2):02d}:{(i % 2) * 30:02d}"
        row   = {"PST_time": pst_t}
        for c, site in site_col_idx.items():
            v = ws.cell(row=r, column=c).value
            try:
                row[f"Req_{site}"] = int(float(str(v).replace(",", "")))
            except:
                row[f"Req_{site}"] = None
        rows.append(row)
    wb.close()
    df = pl.DataFrame(rows)
    print(f"  [UCP] LG Chat req loaded | sites={list(site_col_idx.values())} | {len(df)} intervals")
    return df


def build_html_table(df_pd, title, subtitle="", cases=0, summary="", icon="&#127919;"):
    now = datetime.now()
    ts  = f"Updated {now.strftime('%d-%b-%Y')} · {now.strftime('%I:%M %p')} (VNT)"
    sub = ts + (f" &nbsp;|&nbsp; {subtitle}" if subtitle else "")

    cols    = list(df_pd.columns)
    th_base = 'style="color:#ffffff;padding:7px 10px;border:1px solid #2c4f7c;text-align:center;white-space:nowrap;"'

    headers = []
    for c in cols:
        site = next((s for s in SITES if s in c), None)
        bg   = SITE_HEADER_BG.get(site, "#1e3a5f")
        headers.append(f'<th bgcolor="{bg}" {th_base}>{c}</th>')

    def _td(v, bg="#ffffff", fg="#1a1a1a", bold=False, rowspan=1):
        fw = "font-weight:bold;" if bold else ""
        rs = f' rowspan="{rowspan}"' if rowspan > 1 else ""
        return (
            f'<td{rs} bgcolor="{bg}" '
            f'style="background-color:{bg};color:{fg};{fw}'
            f'padding:5px 10px;border:1px solid #ddd;'
            f'text-align:center;white-space:nowrap;vertical-align:middle;">{v}</td>'
        )

    def _hc_bg(hc_val, req_val):
        try:
            diff = float(hc_val) - float(req_val)
            if diff >= 3:  return "#1B5E20", "#ffffff"
            if diff >= 0:  return "#43A047", "#ffffff"
            if diff >= -3: return "#FB8C00", "#ffffff"
            return "#B71C1C", "#ffffff"
        except:
            return "#ffffff", "#1a1a1a"

    # n_sites computed dynamically per interval (PST+VNT as group key)
    IV_COLS  = ["PST", "VNT", "Forecast (all)", "Productive (all)"]
    divider  = (
        f'<tr><td colspan="{len(cols)}" '
        f'style="padding:1px;background-color:#cfd8dc;border:none;font-size:1px;">&nbsp;</td></tr>'
    )

    rows_html = ""
    i = row_idx = 0

    while i < len(df_pd):
        # Detect group size: rows with same PST+VNT belong to same interval
        cur_key = (str(df_pd.iloc[i].get("PST", "")), str(df_pd.iloc[i].get("VNT", "")))
        n_sites = 1
        while i + n_sites < len(df_pd):
            nxt = df_pd.iloc[i + n_sites]
            if (str(nxt.get("PST", "")), str(nxt.get("VNT", ""))) == cur_key:
                n_sites += 1
            else:
                break
        chunk = df_pd.iloc[i:i + n_sites]
        rbg   = "#f0f4ff" if row_idx % 2 == 0 else "#ffffff"

        for j, (_, row) in enumerate(chunk.iterrows()):
            site  = str(row.get("Site", ""))
            loc_s = LOC_STYLE.get(site, {})
            cells = []

            if j == 0:
                for c in IV_COLS:
                    v = str(row.get(c, "")) if pd.notna(row.get(c)) else "—"
                    if c in ("PST", "VNT"):
                        parts  = v.split(" ") if " " in v else [v]
                        date_p = (f'<span style="font-size:10px;color:#555555;">'
                                  f'{parts[0]}</span><br>') if len(parts) > 1 else ""
                        time_p = f'<b>{parts[-1]}</b>'
                        cells.append(_td(date_p + time_p, bg=rbg, rowspan=n_sites))
                    else:
                        try:    v_f = f"{float(v):.2f}"
                        except: v_f = v
                        cells.append(_td(v_f, bg=rbg, rowspan=n_sites))

            cells.append(_td(site, bg=loc_s.get("bg", "#ffffff"),
                             fg=loc_s.get("fg", "#1a1a1a"), bold=True))

            pa   = row.get("Productive Agents")
            pa_v = str(int(float(pa))) if pd.notna(pa) else "—"
            cells.append(_td(pa_v, bg=rbg, bold=True))

            ph   = row.get("Productive Hours")
            ph_v = f"{float(ph):.2f}" if pd.notna(ph) else "—"
            cells.append(_td(ph_v, bg=rbg))

            hc   = row.get("Heads Contribute")
            req  = row.get("Req Heads")
            # HC có thể là string (e.g. "18.8%") hoặc float → handle cả hai
            if isinstance(hc, str) and hc:
                hc_v = hc          # đã format sẵn, hiển thị plain
            elif pd.notna(hc):
                hc_v = f"{float(hc):.2f}"
            else:
                hc_v = "—"
            # Chỉ tô màu nếu HC là số và có cột Req Heads
            if not isinstance(hc, str) and pd.notna(hc) and pd.notna(req) and hc_v != "—":
                hcb, hcf = _hc_bg(hc, req)
                cells.append(_td(hc_v, bg=hcb, fg=hcf, bold=True))
            else:
                cells.append(_td(hc_v, bg=rbg))

            if "Req Heads" in cols:
                req_v = str(int(float(req))) if pd.notna(req) else "—"
                cells.append(_td(req_v, bg=loc_s.get("bg", "#ffffff"),
                                 fg=loc_s.get("fg", "#1a1a1a"), bold=True))

            rows_html += f"<tr>{''.join(cells)}</tr>"

        rows_html += divider
        i       += n_sites
        row_idx += 1

    sh = (f'  <span style="font-size:11px;color:#444444;">{summary}</span><br>\n'
          if summary else "")
    return (
        f'<p>\n  <b style="color:#006064;font-size:16px;">{icon} {title}</b><br>\n'
        f'  <span style="font-size:12px;color:#555555;">{sub}'
        f' &nbsp;|&nbsp; &#9889; <b>{cases} INTERVALS</b></span><br>\n{sh}</p>\n'
        f'<div style="overflow-x:auto;background-color:#ffffff;padding:2px;">\n'
        f'<table border="1" cellpadding="0" cellspacing="0" bgcolor="#ffffff" '
        f'style="border-collapse:collapse;font-size:12px;'
        f'font-family:Segoe UI,Arial,sans-serif;background-color:#ffffff;">\n'
        f'  <thead><tr>{"".join(headers)}</tr></thead>\n'
        f'  <tbody>{rows_html}</tbody>\n</table>\n</div>'
    )


def send_html_via_webhook(df_pd, title, subtitle="", cases=0, summary="", chunk_size=20, icon="&#127919;", force=False):
    if not force and (df_pd is None or df_pd.empty):
        print(f"[SKIP] '{title}' — empty dataframe")
        return
    n_chunk = max(1, (len(df_pd) + chunk_size - 1) // chunk_size)

    MAX_KB = 26  # Teams webhook limit ~28KB; safe margin

    def _post(payload_str, label):
        try:
            kb = len(payload_str.encode()) / 1024
            if kb > MAX_KB:
                print(f"[WARN] '{label}' payload {kb:.1f}KB > {MAX_KB}KB limit — skipping chunk")
                return
            r  = requests.post(TEAMS_WEBHOOK_URL, headers={"Content-Type": "application/json"},
                               data=payload_str, timeout=30)
            print(f"[OK] '{label}' ({kb:.1f}KB)" if r.status_code in (200, 202)
                  else f"[FAIL] '{label}' [{r.status_code}]: {r.text[:150]}")
        except Exception as e:
            print(f"[ERROR] '{label}': {e}")

    for i in range(n_chunk):
        chunk = df_pd.iloc[i * chunk_size:(i + 1) * chunk_size].reset_index(drop=True)
        label = f"{title} ({i+1}/{n_chunk})" if n_chunk > 1 else title
        html  = build_html_table(
            chunk, label, subtitle,
            len(chunk) if i > 0 else cases,
            summary if i == 0 else "",
            icon=icon,
        )
        _post(json.dumps({"html": html}), label)
        if i < n_chunk - 1:
            time.sleep(1.5)

# %%
now_vnt   = datetime.now(TZ_VNT).replace(tzinfo=None)
base_date = now_vnt.date()
now_pst   = datetime.now(TZ_PST).replace(tzinfo=None)

# ── Load forecast_realtime ─────────────────────────────────────────────────────
print("Loading forecast_realtime...")
fr_raw, fr_exp = load_latest_snapshot(FORECAST_DIR, glob_pattern="In-Progress Interval*")
if fr_raw.is_empty():
    raise RuntimeError("No forecast_realtime files found")

fr_raw = fr_raw.select([c for c in ["Interval Time", "Child Forecast Group",
                                     "Productive Hours", "Productive Agents"]
                         if c in fr_raw.columns])
fr_base = (
    fr_raw
    .with_columns([
        pl.col("Child Forecast Group").replace(CHILD_LOB_MAP, default=None).alias("LOB"),
        pl.col("Productive Hours").cast(pl.Float64, strict=False),
        pl.col("Productive Agents").cast(pl.Float64, strict=False),
    ])
    .filter(pl.col("LOB").is_not_null())
    .with_columns(
        pl.col("Interval Time").cast(pl.Utf8)
          .map_elements(lambda t: _parse_time_str(t, base_date, now_vnt), return_dtype=pl.Datetime)
          .alias("VNT_DT")
    )
    .filter(pl.col("VNT_DT").is_not_null())
    .select(["VNT_DT", "LOB", "Productive Hours", "Productive Agents"])
    .sort(["VNT_DT", "LOB"])
)

all_vnt = fr_base.select("VNT_DT").unique().sort("VNT_DT")
fr_wide = all_vnt
for site in SITES:
    site_df = (
        fr_base.filter(pl.col("LOB") == f"LG Chat {site}")
        .select(["VNT_DT", "Productive Hours", "Productive Agents"])
        .rename({"Productive Agents": f"PA_{site}", "Productive Hours": f"P_{site}"})
    )
    fr_wide = fr_wide.join(site_df, on="VNT_DT", how="left")
print(f"  [FR] Pivoted: {fr_wide.shape}")

# ── Load current_interval ──────────────────────────────────────────────────────
print("Loading current_interval...")
ci_raw, ci_exp = load_latest_snapshot(INTERVAL_DIR, glob_pattern="Current Interval*")
if ci_raw.is_empty():
    raise RuntimeError("No current_interval files found")

ci = (
    ci_raw
    .filter(pl.col("Forecast Group").cast(pl.Utf8).str.contains(INTERVAL_LG_KEY))
    .with_columns([
        pl.col("Interval Time").cast(pl.Utf8)
          .map_elements(lambda t: _parse_time_str(t, base_date, now_vnt), return_dtype=pl.Datetime)
          .alias("VNT_DT"),
        pl.col("Forecasted Hours").cast(pl.Float64, strict=False),
        pl.col("Productive Hours").cast(pl.Float64, strict=False),
    ])
    .filter(pl.col("VNT_DT").is_not_null())
    .select(["VNT_DT", "Forecasted Hours", "Productive Hours"])
    .rename({"Forecasted Hours": "Forecast (all)", "Productive Hours": "Productive (all)"})
    .sort("VNT_DT")
)
print(f"  [CI] Processed: {ci.shape}")

# ── Load UCP Req Heads ─────────────────────────────────────────────────────────
print("Loading UCP...")
try:
    ucp_df = read_ucp_req(UCP_FILE)
except Exception as e:
    print(f"  [WARN] UCP load failed: {e}")
    ucp_df = pl.DataFrame()

# ── Merge & build Current Interval long table ──────────────────────────────────
df = fr_wide.join(ci, on="VNT_DT", how="left")
df = df.with_columns(
    pl.col("VNT_DT")
      .dt.replace_time_zone("Asia/Ho_Chi_Minh")
      .dt.convert_time_zone("America/Los_Angeles")
      .dt.replace_time_zone(None)
      .alias("PST_DT")
).with_columns(pl.col("PST_DT").dt.strftime("%H:%M").alias("PST_time"))

if not ucp_df.is_empty():
    req_cols = [c for c in ucp_df.columns if c.startswith("Req_")]
    df = df.join(ucp_df.select(["PST_time"] + req_cols), on="PST_time", how="left")

for site in SITES:
    if f"P_{site}" in df.columns:
        df = df.with_columns((pl.col(f"P_{site}") * 2).round(2).alias(f"HC_{site}"))

site_cols  = [f"{p}{s}" for s in SITES for p in ["PA_", "P_", "HC_", "Req_"] if f"{p}{s}" in df.columns]
df_final   = df.select(["PST_DT", "VNT_DT", "Forecast (all)", "Productive (all)"] + site_cols).sort("VNT_DT")

rows_long = []
for row in df_final.iter_rows(named=True):
    for site in SITES:
        rows_long.append({
            "PST_DT":            row["PST_DT"],
            "VNT_DT":            row["VNT_DT"],
            "Forecast (all)":    row.get("Forecast (all)"),
            "Productive (all)":  row.get("Productive (all)"),
            "Site":              site,
            "Productive Agents": row.get(f"PA_{site}"),
            "Productive Hours":  row.get(f"P_{site}"),
            "Heads Contribute":  row.get(f"HC_{site}"),
            "Req Heads":         row.get(f"Req_{site}"),
        })

df_long = (
    pl.DataFrame(rows_long).sort("VNT_DT")
    .with_columns([
        pl.col("VNT_DT").dt.strftime("%d-%b %H:%M").alias("VNT"),
        pl.col("PST_DT").dt.strftime("%d-%b %H:%M").alias("PST"),
    ])
    .drop(["VNT_DT", "PST_DT"])
    .select(["PST", "VNT", "Forecast (all)", "Productive (all)",
             "Site", "Productive Agents", "Productive Hours", "Heads Contribute", "Req Heads"])
)
print(f"  [CI] Long table ready: {df_long.shape} | {len(df_long)//len(SITES)} intervals")

# ── Load IC Summary ──────────────────────────────────────────────────────────
# Failure criterion: Interval Compliance == 0 (từ WFM system)
# Forecast(all): từ current_interval (capture_date VNT + interval time)
# Per-site PH: từ forecast_interval_summary (child rows)
print("Loading IC Summary...")
ic_raw = load_all_ic_csv(FORECAST_INTERVAL_DIR)
df_ic  = pl.DataFrame()

if ic_raw.is_empty():
    print("  [IC] ic_raw empty")
else:
    print(f"  [IC] ic_raw: {ic_raw.shape} | columns: {ic_raw.columns}")

    fg_col = (
        "Child Forecast Group" if "Child Forecast Group" in ic_raw.columns else
        "Forecast Group"       if "Forecast Group"       in ic_raw.columns else None
    )
    ic_col = next((c for c in ic_raw.columns if "Interval Compliance" in c), None)
    print(f"  [IC] fg={fg_col!r} | ic_col={ic_col!r}")

    if fg_col is None:
        print("  [IC] ❌ No forecast group column — skip")
    else:
        # ── Step 1: Load ALL rows LG Chat từ forecast_interval_summary ─────
        # IC column mặc định = 0 cho toàn bộ bảng → không dùng để filter
        # Failed criterion: sum(PH all sites) / Forecast(all) < IC_FAIL_THRESHOLD (95%)
        ic_failed = (
            ic_raw
            .filter(pl.col(fg_col).cast(pl.Utf8).str.contains(INTERVAL_LG_KEY))
            .with_columns([
                pl.col("Productive Hours").cast(pl.Float64, strict=False),
                pl.col("Interval Time").cast(pl.Utf8)
                  .map_elements(_parse_ic_time, return_dtype=pl.Datetime)
                  .alias("VNT_DT"),
                pl.col(fg_col).replace(CHILD_LOB_MAP, default=None).alias("LOB"),
            ])
            .filter(pl.col("VNT_DT").is_not_null())
            .with_columns(
                pl.col("VNT_DT")
                  .dt.replace_time_zone("Asia/Ho_Chi_Minh")
                  .dt.convert_time_zone("America/Los_Angeles")
                  .dt.replace_time_zone(None)
                  .alias("PST_DT")
            )
            .filter(pl.col("PST_DT").dt.date()
                      .is_in([now_pst.date(), (now_pst - timedelta(days=1)).date()]))
            .filter(pl.col("VNT_DT") < now_vnt)
        )
        print(f"  [IC] ic_failed (IC=0): {len(ic_failed)} rows")

        if ic_failed.is_empty():
            print("  [IC] No IC=0 rows in date range — all intervals passed")
        else:
            # ── Step 2: Forecast(all) từ current_interval ─────────────────────
            # Dùng capture_date (VNT mtime) + interval_time → tránh cross-day pollution
            _fc: dict = {}
            for fn in sorted(pathlib.Path(INTERVAL_DIR).glob("Current Interval*")):
                if fn.suffix.lower() not in (".csv", ".xlsx"): continue
                try:
                    cap_date = _mtime_dt(fn).date()
                    _df = (pl.read_excel(fn) if fn.suffix.lower() == ".xlsx"
                           else pl.read_csv(fn, infer_schema_length=10000))
                    if _df.is_empty(): continue
                    _fg2 = next((c for c in _df.columns if "Forecast Group" in c), None)
                    if not _fg2: continue
                    for row in _df.filter(
                        pl.col(_fg2).cast(pl.Utf8).str.strip_chars() == INTERVAL_LG_KEY
                    ).iter_rows(named=True):
                        it = str(row.get("Interval Time", "")).strip()
                        fh = row.get("Forecasted Hours")
                        if not it or fh is None: continue
                        try:
                            parts  = it.split(":")
                            h, m   = int(parts[0]) % 24, (int(parts[1]) if len(parts) > 1 else 0)
                            vnt_dt = datetime.combine(cap_date, datetime.min.time()) + timedelta(hours=h, minutes=m)
                            if vnt_dt not in _fc:
                                _fc[vnt_dt] = float(fh)
                        except: pass
                except Exception as e:
                    print(f"  [WARN ci] {fn.name}: {e}")

            fc_df = (
                pl.DataFrame({
                    "VNT_DT": list(_fc.keys()),
                    "Forecast (all)": list(_fc.values()),
                }, schema={"VNT_DT": pl.Datetime("us"), "Forecast (all)": pl.Float64})
                if _fc else
                pl.DataFrame(schema={"VNT_DT": pl.Datetime("us"), "Forecast (all)": pl.Float64})
            )
            print(f"  [IC] FH lookup: {len(fc_df)} intervals from current_interval")

            # ── Step 3: Aggregate per VNT_DT ─────────────────────────────────
            totals = (
                ic_failed.group_by("VNT_DT")
                .agg([
                    pl.col("PST_DT").first().alias("PST_DT"),
                    pl.col("Productive Hours").sum().alias("Productive (all)"),
                ])
                .sort("VNT_DT")
            )
            totals = (
                totals.join(fc_df, on="VNT_DT", how="left")
                if not fc_df.is_empty() else
                totals.with_columns(pl.lit(None).cast(pl.Float64).alias("Forecast (all)"))
            )

            # Failed: sum(PH) < Forecast * 95%
            totals = totals.filter(
                pl.col("Forecast (all)").is_not_null() &
                (pl.col("Forecast (all)") > 0) &
                (pl.col("Productive (all)") < pl.col("Forecast (all)") * IC_FAIL_THRESHOLD)
            )
            failed_vnt_list = sorted(totals["VNT_DT"].to_list())
            print(f"  [IC] {len(failed_vnt_list)} failed intervals (<{IC_FAIL_THRESHOLD*100:.0f}% attainment)")

            # ── Step 4: Per-site rows — chỉ site có PH data ──────────────────
            has_site   = not ic_failed.filter(pl.col("LOB").is_not_null()).is_empty()
            ic_by_site = (
                ic_failed
                .filter(pl.col("LOB").is_not_null())
                .filter(pl.col("Productive Hours").is_not_null())
                .with_columns(pl.col("LOB").str.replace("LG Chat ", "").alias("Site"))
            ) if has_site else pl.DataFrame()

            ic_rows = []
            for vnt_val in failed_vnt_list:
                tr           = totals.filter(pl.col("VNT_DT") == vnt_val).row(0, named=True)
                pst_dt_val   = tr["PST_DT"]
                fc_all       = tr.get("Forecast (all)")
                prod_all     = tr["Productive (all)"]
                pst_time_str = pst_dt_val.strftime("%H:%M") if pst_dt_val else None

                if has_site and not ic_by_site.is_empty():
                    for r in ic_by_site.filter(pl.col("VNT_DT") == vnt_val).iter_rows(named=True):
                        site = r["Site"]
                        ph   = r.get("Productive Hours")
                        if ph is None: continue
                        pa  = round(float(ph) * 2, 2)
                        hc  = f"{float(ph)/float(fc_all)*100:.1f}%" if fc_all and float(fc_all) > 0 else None
                        req_h = None
                        if not ucp_df.is_empty() and pst_time_str:
                            rr = ucp_df.filter(pl.col("PST_time") == pst_time_str)
                            if not rr.is_empty() and f"Req_{site}" in rr.columns:
                                req_h = rr[f"Req_{site}"][0]
                        ic_rows.append({
                            "PST_DT": pst_dt_val, "VNT_DT": vnt_val,
                            "Forecast (all)": fc_all, "Productive (all)": prod_all,
                            "Site": site, "Productive Agents": pa,
                            "Productive Hours": ph, "Heads Contribute": hc,
                        })
                else:
                    ph  = prod_all
                    pa  = round(float(ph) * 2, 2) if ph else None
                    hc  = f"{float(ph)/float(fc_all)*100:.1f}%" if ph and fc_all and float(fc_all) > 0 else None
                    ic_rows.append({
                        "PST_DT": pst_dt_val, "VNT_DT": vnt_val,
                        "Forecast (all)": fc_all, "Productive (all)": prod_all,
                        "Site": "All", "Productive Agents": pa,
                        "Productive Hours": ph, "Heads Contribute": hc,
                    })

            print(f"  [IC] ic_rows: {len(ic_rows)}")
            if ic_rows:
                df_ic = (
                    pl.DataFrame(ic_rows, infer_schema_length=None,
                                 schema_overrides={
                                     "VNT_DT": pl.Datetime, "PST_DT": pl.Datetime,
                                     "Forecast (all)": pl.Float64, "Productive (all)": pl.Float64,
                                     "Productive Agents": pl.Float64, "Productive Hours": pl.Float64,
                                     "Heads Contribute": pl.Utf8,
                                 })
                    .sort("VNT_DT")
                    .with_columns([
                        pl.col("VNT_DT").dt.strftime("%d-%b %H:%M").alias("VNT"),
                        pl.col("PST_DT").dt.strftime("%d-%b %H:%M").alias("PST"),
                    ])
                    .drop(["VNT_DT", "PST_DT"])
                    .select(["PST", "VNT", "Forecast (all)", "Productive (all)",
                             "Site", "Productive Agents", "Productive Hours",
                             "Heads Contribute"])
                )
                print(f"  [IC] df_ic ready: {df_ic.shape}")

print("\nAll tables built successfully")

# %%
# ── Current Interval ───────────────────────────────────────────────────────────

df_ci_pd = df_long.to_pandas()
n_ci     = len(df_ci_pd) // len(SITES)
summary_ci = "  |  ".join(
    f"{s}: {int(df_ci_pd[df_ci_pd['Site']==s]['Productive Agents'].dropna().iloc[-1])} agents"
    for s in SITES
    if not df_ci_pd[df_ci_pd["Site"]==s]["Productive Agents"].dropna().empty
)
send_html_via_webhook(
    df_ci_pd,
    title      = "LG Chat — Current Interval",
    subtitle   = f"In-progress intervals — {now_vnt.strftime('%d-%b %H:%M')} VNT",
    cases      = n_ci,
    summary    = summary_ci,
    chunk_size = len(SITES) * 2,   # ~8 rows/chunk → safe for Teams
    icon       = "🔴"
)

# ── IC Summary: Failed Intervals ───────────────────────────────────────────────
pst_today = now_pst.date()
# ── DEBUG: xác nhận trạng thái df_ic ────────────────────────────────────────
print(f"[DEBUG Cell3] df_ic type      : {type(df_ic)}")
print(f"[DEBUG Cell3] df_ic.is_empty(): {df_ic.is_empty() if hasattr(df_ic, 'is_empty') else 'N/A'}")
print(f"[DEBUG Cell3] df_ic.shape     : {df_ic.shape if hasattr(df_ic, 'shape') else 'N/A'}")
if hasattr(df_ic, 'is_empty') and not df_ic.is_empty():
    print(df_ic.head(6))
# ─────────────────────────────────────────────────────────────────────────────
if df_ic is not None and not df_ic.is_empty():
    df_ic_pd = df_ic.to_pandas()
    # Đếm unique intervals (PST+VNT) thay vì chia cho SITES
    n_ic     = df_ic.select(["PST", "VNT"]).unique().height
    ic_title   = f"LG Chat — IC Summary | ⚠️ {n_ic} Failed Interval(s) | PST {(now_pst - timedelta(days=1)).strftime('%d-%b')} – {now_pst.strftime('%d-%b')}"
    ic_summary = f"⚠️ {n_ic} failed interval(s) detected across {len(SITES)} sites"
else:
    # Không có failed interval → gửi bảng trống kèm note
    df_ic_pd = pd.DataFrame(columns=["PST", "VNT", "Forecast (all)", "Productive (all)",
                                      "Site", "Productive Agents", "Productive Hours",
                                      "Heads Contribute", "Req Heads"])
    n_ic     = 0
    ic_title   = "LG Chat — IC Summary | ✅ All Intervals Passed"
    ic_summary = f"All completed intervals >= {IC_FAIL_THRESHOLD*100:.0f}% attainment — no failed intervals"

# ── DEBUG: trạng thái df_ic trước khi gửi ───────────────────────────────────
print(f"[DEBUG] df_ic type      : {type(df_ic)}")
print(f"[DEBUG] df_ic is None   : {df_ic is None}")
if df_ic is not None:
    print(f"[DEBUG] df_ic.is_empty(): {df_ic.is_empty()}")
    print(f"[DEBUG] df_ic.shape     : {df_ic.shape}")
print(f"[DEBUG] df_ic_pd.empty  : {df_ic_pd.empty}")
print(f"[DEBUG] df_ic_pd.shape  : {df_ic_pd.shape}")
print(f"[DEBUG] n_ic            : {n_ic}")
print(f"[DEBUG] ic_title        : {ic_title}")
# ─────────────────────────────────────────────────────────────────────────────
send_html_via_webhook(
    df_ic_pd,
    title      = ic_title,
    subtitle   = (
        f"Past completed intervals · "
        f"PST {pst_today - timedelta(days=1)} → {pst_today}"
    ),
    cases      = n_ic,
    summary    = ic_summary,
    chunk_size = len(SITES) * 2,   # ~8 rows/chunk → safe for Teams
    force      = True,
)
print(f"[IC] Done | {n_ic} failed intervals")

print(f"\nDone | CI={len(df_long)//len(SITES)} intervals | IC={len(df_ic)//len(SITES) if not df_ic.is_empty() else 0} failed")


