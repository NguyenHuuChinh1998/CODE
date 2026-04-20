# %%
import os
import io
import time
from datetime import datetime

from PIL import Image
import win32clipboard

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import TimeoutException, NoSuchElementException, NoAlertPresentException

from webdriver_manager.chrome import ChromeDriverManager
from urllib.parse import urlparse

# %%
expedia_url = "https://cnxnice02b.nicecloudsvc.com/wfm/supervisor/reports-generate"
teams_url = "https://teams.microsoft.com/"
group_chat_name = "[BOT] IC Monitoring"
parsed_url = urlparse(expedia_url)
path_fragment = parsed_url.path.split('/')[-1]
report_name = "IEX"
wait_seconds = 10
start_time = datetime.now()

chrome_options = Options()
chrome_options.add_argument(r'--user-data-dir=C:/temp/new_chrome_profile')
chrome_options.add_argument(r'--profile-directory=Default')
chrome_options.add_argument("--start-maximized")

service = Service(ChromeDriverManager().install())

service = Service(r"C:\Users\huuchinh.nguyen\Concentrix Corporation\WFM-Expedia-HCM - Branding files\Rawdata\CAPTURE\chromedriver-win64\chromedriver.exe") 
driver = webdriver.Chrome(service=service, options=chrome_options)

    

wait = WebDriverWait(driver, 15)
driver.get(expedia_url)
    
def check_and_login(driver, expedia_url, wait_time=10):
    driver.get(expedia_url)
    time.sleep(10)  # Let the page load

    try:
        sign_in_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-testid="console-okta-sign-in"]'))
        )
        print("🔑 Sign-in required detected! Clicking...")
        sign_in_button.click()
        time.sleep(2)
        try:
            keep_signed_in_label = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'label[for="input36"][data-se-for-name="rememberMe"]'))
            )
            keep_signed_in_label.click()
            time.sleep(1)
        except TimeoutException:
            print("No 'Keep me signed in' option found. Skipping.")
        try:
            next_button = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, 'input.button.button-primary[type="submit"][value="Next"]'))
            )
            next_button.click()
            time.sleep(10)
        except TimeoutException:
            print("No 'Next' button found. Skipping.")
        print("🎉 Login successful! Reloading the page...")
        driver.get(expedia_url)
    except TimeoutException:
        print("✅ No sign-in required. Continuing with expedia_url...")

check_and_login(driver, expedia_url)

time.sleep(5)

try:
    body = wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
    if body.text.strip() == "Not Found":
        print("Page Not Found. Reloading...")
        driver.refresh()
        wait.until(EC.presence_of_element_located((By.CLASS_NAME, "legacy-wrapper")))
except TimeoutException:
    print("Timeout waiting for body tag")

time.sleep(5)

# %%
wait = WebDriverWait(driver, 10)
max_retries = 5
retry_count = 0
found_and_displayed = False

while retry_count < max_retries:
    try:
        # Wait for the element to appear on the DOM
        generate_link = wait.until(EC.presence_of_element_located((By.XPATH, '//a[@title="Generate" and contains(@class, "sub-menu-item")]')))

        if generate_link.is_displayed():
            print(f"Attempt {retry_count+1}: 'Generate' link is visible.")
            found_and_displayed = True
            break  # Exit the loop if the link is found and visible
        else:
            print(f"Attempt {retry_count+1}: 'Generate' link is not visible. Waiting 5s and reopening the page...")
            time.sleep(10)
            driver.execute_script("window.open('https://cnxnice02b.nicecloudsvc.com/wfm/supervisor/reports-generate', '_blank');")
            driver.switch_to.window(driver.window_handles[-1])
    except:
        print(f"Attempt {retry_count+1}: 'Generate' link not found. Waiting 5s and reopening the page...")
        time.sleep(10)
        driver.execute_script("window.open('https://cnxnice02b.nicecloudsvc.com/wfm/supervisor/reports-generate', '_blank');")
        driver.switch_to.window(driver.window_handles[-1])

    retry_count += 1

if not found_and_displayed:
    print("Tried 5 times but 'Generate' link was not found or not visible. Closing the program.")
    driver.quit()

# %%
driver.get(expedia_url)
time.sleep(5)
driver.get("https://cnxnice02b.nicecloudsvc.com/supv/reportAction.mvc?schRptOid=8aa89bca8b4b614b018d4b88862d476c")
time.sleep(5)

# %%
import datetime
import time

try:
    wait = WebDriverWait(driver, 10)
    input_date = wait.until(EC.presence_of_element_located((By.ID, "stAbsDate")))
    input_date.clear()
    input_date.send_keys("5/26/25")
    time.sleep(2)
    input_end_date = wait.until(EC.presence_of_element_located((By.ID, "endAbsDate")))
    input_end_date.clear()
    input_end_date.send_keys("5/26/25")
    time.sleep(2)
    generate_button = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@type='submit' and @value='Generate']")))
    generate_button.click()
    
    time.sleep(5)  # hoặc dùng wait đến khi alert có mặt

    try:
        alert = driver.switch_to.alert
        alert.accept()  # Bấm nút OK trên alert
    except NoAlertPresentException:
        pass
except TimeoutException:
    pass


# %%
time.sleep(60)

driver.get("https://cnxnice02b.nicecloudsvc.com/wfm/supervisor/reports-view")

wait = WebDriverWait(driver, 10)

iframe = wait.until(EC.presence_of_element_located((By.CLASS_NAME, "legacy-wrapper")))
driver.switch_to.frame(iframe)

refresh_button = wait.until(EC.element_to_be_clickable((By.XPATH, '//input[@type="submit" and @value="Refresh"]')))
refresh_button.click()

time.sleep(10)

link = wait.until(EC.element_to_be_clickable((By.LINK_TEXT, "Agent Schedules")))
link.click()

driver.switch_to.default_content()

# %%
time.sleep(10)

driver.get(teams_url)

import datetime
import time

try:
    wait = WebDriverWait(driver, 10)
    group_chat = wait.until(EC.presence_of_element_located((By.XPATH, f"//span[contains(text(),'{group_chat_name}')]")))
    group_chat.click()
except TimeoutException:
    pass

time.sleep(3)

chat_box = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "div[role='textbox']")))
chat_box.click()

now = datetime.datetime.now()
realtime = now.strftime("%I:%M %p")  
chat_text = f"{report_name} downloaded at {realtime} (VNT)"

actions = ActionChains(driver)
actions.send_keys(chat_text)
actions.perform()

time.sleep(2)

actions.send_keys(Keys.ENTER).perform()

time.sleep(10)

driver.quit()

from openpyxl import Workbook, load_workbook
from datetime import datetime

end_time = datetime.now()
execution_time = (end_time - start_time).total_seconds()

log_file = "bot_log/bot_capture_log.xlsx"

if not os.path.exists(log_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Log"
    ws.append(["BOT Name", "Run at", "End at", "Execution Time (seconds)"])
else:
    wb = load_workbook(log_file)
    ws = wb["Log"]

ws.append([
    report_name,
    start_time.strftime('%Y-%m-%d %H:%M:%S'),
    end_time.strftime('%Y-%m-%d %H:%M:%S'),
    execution_time
])

wb.save(log_file)

